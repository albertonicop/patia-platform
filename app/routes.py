# -*- coding: utf-8 -*-
import resend
from email_validator import validate_email, EmailNotValidError
import string
from datetime import datetime, timedelta
from io import BytesIO
import stripe
import secrets
import uuid
import re
import json
from decimal import Decimal, InvalidOperation
from flask import Blueprint, render_template, request, redirect, url_for, flash as flask_flash, session, current_app, send_file, jsonify, abort, has_request_context, g
from flask_babel import force_locale, gettext
from sqlalchemy import case, exists, func, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload
from urllib.parse import urljoin, urlparse
from . import csrf, db, limiter
from .barcodes import (
    automatic_sku,
    find_company_product_by_barcode,
    lookup_barcode,
    normalize_barcode,
)
from .models import (
    CashRegisterSession,
    Customer,
    CustomerCreditMovement,
    InventoryMovement,
    InventoryRestockEvent,
    MonthlyOwnerReport,
    Organization,
    OrganizationInvitation,
    OrganizationMember,
    Product,
    Sale,
    SalesTicket,
    StripeWebhookEvent,
    Supplier,
    User,
)
from .cash.services import (
    expected_cash as cash_expected_amount,
    open_cash_session,
    record_cash_movement,
)
from .inventory.services import (
    record_inventory_movement,
    record_opening_balance,
)
from .inventory.imports import (
    MAX_IMPORT_ROWS,
    apply_catalog,
    inspect_catalog,
)
from .credit.services import (
    CreditError,
    CreditLimitExceeded,
    CreditNotEnabled,
    record_credit_charge,
    record_credit_reversal,
)
from .money import MONEY_ZERO, money_decimal, money_json, money_sum
from .currencies import (
    COUNTRY_OPTIONS,
    SUPPORTED_CURRENCIES,
    country_defaults,
    format_currency as format_organization_currency,
    normalize_country_code,
    normalize_currency_code,
    normalize_locale_code,
    organization_money_context,
)
from .timezones import (
    DEFAULT_TIMEZONE,
    TIMEZONE_CHOICES,
    local_date_bounds_utc,
    local_today,
    safe_timezone_name,
    utc_to_local,
)
from .team.services import (
    active_membership,
    authentication_required_response,
    ensure_owner_organization,
    has_permission,
    membership_for_login,
    organization_owner,
    require_permission,
    require_roles,
)

main = Blueprint("main", __name__)
SUPPORTED_LANGUAGES = {"es", "en"}


def flash(message, category="message"):
    """Translate application-owned notices without altering user-provided data."""
    return flask_flash(gettext(message), category)


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    user = User.query.get(user_id)
    if not user:
        language = session.get("language", "es")
        session.clear()
        session["language"] = language if language in SUPPORTED_LANGUAGES else "es"
        session["session_expired"] = True
        return None
    if user.session_token and user.session_token != session.get("session_token"):
        session.clear()
        session["kicked_out"] = True
        return None
    membership = membership_for_login(user)
    if membership is None:
        session.clear()
        session["membership_disabled"] = True
        return None
    session["organization_id"] = membership.organization_id
    sync_user_plan(membership.organization.owner)
    return user


def current_organization_id(user):
    """Resolve tenant membership without trusting an organization from input."""
    if user is None:
        return None
    membership = active_membership(user)
    if membership is None:
        if not has_request_context():
            return None
        membership = membership_for_login(user)
        if membership is None:
            return None
        session["organization_id"] = membership.organization_id
    return membership.organization_id


def current_organization_owner(user):
    """Billing/legacy owner for the active tenant."""
    if user is None:
        return None
    return organization_owner(user) or user


def _safe_next_url(candidate):
    if not candidate:
        return url_for("main.dashboard")
    base = urlparse(request.host_url)
    target = urlparse(urljoin(request.host_url, candidate))
    if target.scheme in {"http", "https"} and target.netloc == base.netloc:
        return target.path + (f"?{target.query}" if target.query else "")
    return url_for("main.dashboard")


@main.route("/language", methods=["POST"])
def set_language():
    language = request.form.get("language", "").lower()
    if language not in SUPPORTED_LANGUAGES:
        language = "es"
    session["language"] = language
    user = current_user()
    if user and user.preferred_language != language:
        user.preferred_language = language
        db.session.commit()
    return redirect(_safe_next_url(request.form.get("next") or request.referrer))


def trial_expired(user):
    if not user:
        return True
    access_user = current_organization_owner(user)
    if has_pro_access(access_user):
        return False
    days_used = (datetime.utcnow() - access_user.created_at).days
    return days_used >= 14


def _trial_access_response(user, *, json_response=False):
    if not trial_expired(user):
        return None
    if json_response:
        return jsonify({
            "ok": False,
            "error": gettext("Tu periodo de prueba terminó. Activa PATIA Pro para continuar."),
        }), 403
    flash(
        "Tu periodo de prueba terminó. Activa PATIA Pro para continuar.",
        "danger",
    )
    return render_template("trial_expired.html"), 403


def money(value, currency_code=None, locale_code=None):
    cached_context = getattr(g, "_organization_money_context", None) if has_request_context() else None
    organization = None
    if cached_context:
        active_currency, active_locale = cached_context
    elif has_request_context():
        user = current_user()
        membership = active_membership(user) if user else None
        organization = membership.organization if membership else None
        active_currency, active_locale = organization_money_context(organization)
        g._organization_money_context = (active_currency, active_locale)
    else:
        active_currency, active_locale = organization_money_context(None)
    return format_organization_currency(
        value,
        currency_code or active_currency,
        locale_code or active_locale,
    )


def _sale_ticket_key(sale):
    """Identificador interno estable, incluso para ventas históricas sin UUID."""
    return (
        sale.sales_ticket.public_id
        if sale.sales_ticket
        else sale.ticket_id or f"sale-{sale.id}"
    )


def _short_sale_folio(sales):
    """Folio legible y estable sin reemplazar el identificador interno."""
    if sales[0].sales_ticket:
        return sales[0].sales_ticket.folio
    ticket_id = sales[0].ticket_id
    if ticket_id:
        try:
            folio_number = uuid.UUID(ticket_id).int % 1_000_000
            return f"V-{folio_number:06d}"
        except (ValueError, TypeError, AttributeError):
            pass
    return f"V-{min(sale.id for sale in sales):06d}"


def _create_sales_ticket(actor, payment_method, *, ticket_id=None):
    """Allocate a per-company ticket number atomically inside the sale transaction."""
    owner = current_organization_owner(actor)
    organization_id = current_organization_id(actor)
    next_number = db.session.execute(
        update(User)
        .where(User.id == owner.id)
        .values(next_ticket_number=User.next_ticket_number + 1)
        .returning(User.next_ticket_number)
        .execution_options(synchronize_session=False)
    ).scalar_one()
    organization = active_membership(actor).organization
    currency_code, locale_code = organization_money_context(organization)
    ticket = SalesTicket(
        organization_id=organization_id,
        user_id=owner.id,
        number=next_number - 1,
        public_id=ticket_id or str(uuid.uuid4()),
        payment_method=payment_method,
        currency_code=currency_code,
        locale_code=locale_code,
        cashier_member_id=active_membership(actor).id,
    )
    db.session.add(ticket)
    db.session.flush()
    return ticket


PAYMENT_METHOD_LABELS = {
    "cash": "Efectivo",
    "card": "Tarjeta",
    "transfer": "Transferencia",
    "other": "Otro",
    "credit": "Crédito",
}


def _payment_method_label(value):
    return gettext(PAYMENT_METHOD_LABELS.get(value, "No especificado"))


def _selected_customer(organization_id, raw_customer_id):
    if raw_customer_id in (None, ""):
        return None
    try:
        customer_id = int(raw_customer_id)
    except (TypeError, ValueError) as exc:
        raise ValueError(gettext("Selecciona un cliente válido.")) from exc
    customer = Customer.query.filter_by(
        id=customer_id,
        organization_id=organization_id,
        is_active=True,
    ).first()
    if not customer:
        raise ValueError(
            gettext("El cliente seleccionado no está disponible.")
        )
    return customer


def _credit_override_rate_limit_key():
    return ":".join(
        (
            request.remote_addr or "unknown",
            str(session.get("user_id") or "anonymous"),
            str(session.get("organization_id") or "no-organization"),
        )
    )


def _credit_override_rate_limit_exempt():
    payload = request.get_json(silent=True)
    return not (
        isinstance(payload, dict)
        and payload.get("payment_method") == "credit"
        and payload.get("credit_override")
        and str(payload.get("override_pin") or "").strip()
    )


def _credit_override_rate_limit_response(_request_limit):
    response = jsonify(
        {
            "ok": False,
            "error": gettext(
                "Demasiados intentos de autorización. Espera un minuto."
            ),
            "error_code": "credit_override_rate_limited",
        }
    )
    response.status_code = 429
    return response


def _ticket_business_value(value):
    """Hide empty and known placeholder business data from customer tickets."""
    cleaned = str(value or "").strip()
    normalized = re.sub(r"\s+", " ", cleaned).casefold()
    if normalized in {
        "",
        "-",
        "n/a",
        "na",
        "no configurado",
        "no configurada",
        "dirección no configurada",
        "direccion no configurada",
    }:
        return ""
    compact = re.sub(r"\D", "", cleaned)
    if compact and set(compact) == {"0"}:
        return ""
    return cleaned


def _subscription_status_label(value):
    labels = {
        "active": gettext("Activa"),
        "trialing": gettext("En prueba"),
        "past_due": gettext("Pago pendiente"),
        "unpaid": gettext("Sin pagar"),
        "canceled": gettext("Cancelada"),
        "incomplete": gettext("Incompleta"),
        "incomplete_expired": gettext("Incompleta y vencida"),
        "paused": gettext("Pausada"),
    }
    return labels.get((value or "").lower(), gettext("Sin suscripción"))


def _translated_payment_method_labels():
    return {key: gettext(label) for key, label in PAYMENT_METHOD_LABELS.items()}


def _translated_timezone_choices():
    labels = {
        "America/Mexico_City": gettext("Ciudad de México"),
        "America/Cancun": gettext("Cancún"),
        "America/Tijuana": gettext("Tijuana"),
        "America/Hermosillo": gettext("Hermosillo"),
        "America/Chihuahua": gettext("Chihuahua"),
    }
    return [
        (timezone_name, labels[timezone_name])
        for timezone_name, _label in TIMEZONE_CHOICES
    ]


def _group_sales_by_ticket(sales, *, limit=None, timezone_name=DEFAULT_TIMEZONE):
    grouped = {}
    for sale in sales:
        key = _sale_ticket_key(sale)
        group = grouped.setdefault(key, {
            "ticket_id": key,
            "ticket": sale.sales_ticket,
            "sales": [],
            "created_at": sale.created_at,
            "total": 0,
            "item_count": 0,
            "payment_method": (
                sale.sales_ticket.payment_method
                if sale.sales_ticket
                else sale.payment_method
            ),
        })
        group["sales"].append(sale)
        group["total"] += sale.total
        group["item_count"] += sale.quantity
        if sale.created_at < group["created_at"]:
            group["created_at"] = sale.created_at
    result = list(grouped.values())
    for group in result:
        group["folio"] = _short_sale_folio(group["sales"])
        group["created_at_local"] = utc_to_local(
            group["created_at"],
            timezone_name,
        )
    result.sort(key=lambda group: group["created_at"], reverse=True)
    return result[:limit] if limit else result


MANAGED_SUBSCRIPTION_STATUSES = {"active", "trialing", "past_due", "unpaid"}
KNOWN_SUBSCRIPTION_STATUSES = {
    "active", "trialing", "past_due", "unpaid", "canceled",
    "incomplete", "incomplete_expired", "paused",
}


class StripeEventIgnored(Exception):
    """Evento válido de Stripe que no pertenece a esta integración."""


def _as_utc_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.utcfromtimestamp(int(value))


def has_pro_access(user, now=None):
    from .plans import subscription_access_is_active

    return subscription_access_is_active(
        user,
        now=now,
        grace_days=current_app.config.get(
            "STRIPE_PAST_DUE_GRACE_DAYS", 3
        ),
    )


def sync_user_plan(user, now=None):
    user.plan = "pro" if has_pro_access(user, now=now) else "trial"


def _public_url(path):
    return f"{current_app.config['PUBLIC_BASE_URL']}{path}"


def _subscription_has_configured_price(subscription):
    from .plans import configured_price_plan

    items = (subscription.get("items") or {}).get("data") or []
    return any(
        configured_price_plan(
            current_app.config, (item.get("price") or {}).get("id")
        )
        for item in items
    )


def _subscription_plan_code(subscription, *, preserve_legacy=False):
    from .plans import (
        GRANDFATHERED,
        PAID_PLAN_CODES,
        configured_price_plan,
        normalize_plan_code,
    )

    metadata_code = normalize_plan_code(
        (subscription.get("metadata") or {}).get("plan_code"),
        default="",
    )
    items = (subscription.get("items") or {}).get("data") or []
    plans = {
        configured_price_plan(
            current_app.config, (item.get("price") or {}).get("id")
        )
        for item in items
    }
    plans.discard(None)
    if len(plans) != 1:
        raise StripeEventIgnored(
            "La suscripción no corresponde a un único plan configurado."
        )
    price_plan = plans.pop()
    if metadata_code in PAID_PLAN_CODES and metadata_code != price_plan:
        raise StripeEventIgnored(
            "El plan de los metadatos no coincide con el Price ID."
        )
    if metadata_code in PAID_PLAN_CODES:
        return metadata_code
    if preserve_legacy:
        return GRANDFATHERED
    return price_plan


def _subscription_period_end(subscription):
    value = subscription.get("current_period_end")
    if value:
        return _as_utc_datetime(value)
    periods = [
        item.get("current_period_end")
        for item in ((subscription.get("items") or {}).get("data") or [])
        if item.get("current_period_end")
    ]
    return _as_utc_datetime(max(periods)) if periods else None


def _invoice_subscription_id(invoice):
    subscription_id = invoice.get("subscription")
    if subscription_id:
        return subscription_id
    details = (invoice.get("parent") or {}).get("subscription_details") or {}
    return details.get("subscription")


def _event_is_newer(user, stripe_created_at, event_family):
    if event_family == "invoice":
        watermark = user.stripe_invoice_updated_at
    elif event_family == "subscription":
        watermark = user.stripe_subscription_updated_at
    else:
        raise ValueError("Familia de evento Stripe no soportada.")
    return not watermark or stripe_created_at >= watermark


def _validate_subscription(subscription, user=None, customer_id=None):
    if not _subscription_has_configured_price(subscription):
        raise StripeEventIgnored(
            "La suscripción no usa un Price ID configurado en PATIA."
        )
    subscription_customer = subscription.get("customer")
    if customer_id and subscription_customer != customer_id:
        raise StripeEventIgnored("El cliente de Stripe no coincide.")
    if user and user.stripe_customer_id and subscription_customer != user.stripe_customer_id:
        raise StripeEventIgnored("La suscripción no pertenece al cliente guardado.")


def _find_subscription_user(subscription):
    subscription_id = subscription.get("id")
    user = User.query.filter_by(stripe_subscription_id=subscription_id).first()
    if user:
        return user
    user_id = (subscription.get("metadata") or {}).get("user_id")
    if user_id and str(user_id).isdigit():
        return db.session.get(User, int(user_id))
    return None


def _ensure_stripe_ids_available(user, customer_id, subscription_id):
    customer_owner = User.query.filter(
        User.stripe_customer_id == customer_id,
        User.id != user.id,
    ).first()
    subscription_owner = User.query.filter(
        User.stripe_subscription_id == subscription_id,
        User.id != user.id,
    ).first()
    if customer_owner or subscription_owner:
        raise StripeEventIgnored("Identificador Stripe ya vinculado a otro usuario.")


def _sync_subscription_state(user, subscription, stripe_created_at, deleted=False):
    if not _event_is_newer(user, stripe_created_at, "subscription"):
        return False
    status = "canceled" if deleted else (subscription.get("status") or "").lower()
    if status not in KNOWN_SUBSCRIPTION_STATUSES:
        raise StripeEventIgnored(f"Estado de suscripción no soportado: {status}")
    user.stripe_customer_id = subscription.get("customer") or user.stripe_customer_id
    user.stripe_subscription_id = subscription.get("id") or user.stripe_subscription_id
    user.subscription_status = status
    user.current_period_end = _subscription_period_end(subscription) or user.current_period_end
    user.cancel_at_period_end = bool(subscription.get("cancel_at_period_end", False))
    user.stripe_subscription_updated_at = stripe_created_at
    if not deleted:
        metadata = subscription.get("metadata") or {}
        resolved_plan = _subscription_plan_code(
            subscription,
            preserve_legacy=not (
                user.subscription_plan_code or metadata.get("plan_code")
            ),
        )
        user.subscription_plan_code = resolved_plan
        if user.pending_plan_code == resolved_plan:
            user.pending_plan_code = None
            user.pending_plan_effective_at = None
    if status != "past_due":
        user.next_payment_attempt = None
    sync_user_plan(user)
    return True


def _has_managed_stripe_subscription(user):
    if not user.stripe_subscription_id:
        return False
    if user.subscription_status in MANAGED_SUBSCRIPTION_STATUSES:
        return True
    if current_app.config["STRIPE_DISABLED"]:
        return True
    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    try:
        subscription = stripe.Subscription.retrieve(user.stripe_subscription_id)
        _validate_subscription(subscription, user=user)
        _sync_subscription_state(user, subscription, datetime.utcnow())
        db.session.commit()
        return subscription.get("status") in MANAGED_SUBSCRIPTION_STATUSES
    except StripeEventIgnored:
        db.session.rollback()
        return False
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "No se pudo comprobar la suscripción antes de eliminar al usuario"
        )
        return True


def send_email(
    to,
    subject,
    html,
    language="es",
    *,
    idempotency_key=None,
):
    api_key = current_app.config.get("RESEND_API_KEY")
    sender = current_app.config.get("RESEND_FROM")
    if not api_key or not sender:
        current_app.logger.error("Resend no está configurado; correo no enviado.")
        return False
    try:
        with force_locale(language if language in SUPPORTED_LANGUAGES else "es"):
            resend.api_key = api_key
            params = {
                "from": sender,
                "to": to,
                "subject": gettext(subject),
                "html": html,
            }
            options = (
                {"idempotency_key": idempotency_key}
                if idempotency_key
                else None
            )
            if options:
                resend.Emails.send(params, options=options)
            else:
                resend.Emails.send(params)
        return True
    except Exception:
        current_app.logger.exception("No se pudo enviar un correo con Resend.")
        return False


@main.route("/register", methods=["GET", "POST"])
@limiter.limit("3 per hour", methods=["POST"])
def register():
    if request.method == "POST":
        selected_language = session.get("language", "es")
        if selected_language not in SUPPORTED_LANGUAGES:
            selected_language = "es"
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        required_fields = {
            "first_name": "nombre",
            "last_name": "apellido",
            "company_name": "nombre de la empresa",
            "phone": "teléfono",
            "address": "dirección",
            "city": "ciudad",
            "state": "estado",
            "business_type": "giro del negocio",
            "postal_code": "código postal",
            "email": "correo",
            "password": "contraseña",
        }
        missing = [
            label
            for field, label in required_fields.items()
            if not request.form.get(field, "").strip()
        ]
        if missing:
            flash(gettext("Completa todos los campos obligatorios."), "danger")
            return render_template(
                "auth.html",
                title=gettext("Crear cuenta"),
                button=gettext("Crear cuenta"),
                mode="register",
                plan=request.form.get("plan"),
                form_data=request.form,
            ), 400
        try:
            validate_email(email, check_deliverability=True)
        except EmailNotValidError:
            flash(gettext("El correo no es válido o no existe."), "danger")
            return render_template(
                "auth.html",
                title=gettext("Crear cuenta"),
                button=gettext("Crear cuenta"),
                mode="register",
                plan=request.form.get("plan"),
                form_data=request.form,
            ), 400

        if len(password) < 8:
            flash(gettext("La contraseña debe tener al menos 8 caracteres."), "danger")
            return render_template(
                "auth.html",
                title=gettext("Crear cuenta"),
                button=gettext("Crear cuenta"),
                mode="register",
                plan=request.form.get("plan"),
                form_data=request.form,
            ), 400
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        company_name = request.form.get("company_name", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        state = request.form.get("state", "").strip()
        business_type = request.form.get("business_type", "").strip()
        postal_code = request.form.get("postal_code", "").strip()

        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            flash(gettext("Ese correo ya está registrado."), "danger")
            return render_template(
                "auth.html",
                title=gettext("Crear cuenta"),
                button=gettext("Crear cuenta"),
                mode="register",
                plan=request.form.get("plan"),
                form_data=request.form,
            ), 409

        from .plans import PRO, STARTER

        requested_plan = str(
            request.form.get("plan")
            or request.args.get("plan")
            or STARTER
        ).upper()
        if requested_plan not in {STARTER, PRO}:
            requested_plan = STARTER
        user = User(
            email=email,
            company_name=company_name,
            trial_plan_code=requested_plan,
        )
        country_code = normalize_country_code(request.form.get("country_code"))
        currency_code = normalize_currency_code(request.form.get("currency_code"))
        suggested_currency, suggested_locale = country_defaults(country_code)
        if not request.form.get("currency_code"):
            currency_code = suggested_currency
        locale_code = normalize_locale_code(
            request.form.get("locale_code") or suggested_locale,
            currency_code,
        )
        user.preferred_language = selected_language
        user.first_name = first_name
        user.last_name = last_name
        user.phone = phone
        user.address = address
        user.city = city
        user.state = state
        user.business_type = business_type
        user.postal_code = postal_code
        user.set_password(password)

        db.session.add(user)
        try:
            db.session.flush()
            membership = ensure_owner_organization(user)
            membership.organization.country_code = country_code
            membership.organization.currency_code = currency_code
            membership.organization.locale_code = locale_code
            membership.organization.currency = currency_code
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(gettext("Ese correo ya está registrado."), "danger")
            return render_template(
                "auth.html",
                title=gettext("Crear cuenta"),
                button=gettext("Crear cuenta"),
                mode="register",
                plan=request.form.get("plan"),
                form_data=request.form,
            ), 409

        session.clear()
        session["language"] = user.preferred_language
        session["user_id"] = user.id
        session["organization_id"] = membership.organization_id
        session["post_verify_destination"] = (
            "subscribe"
            if requested_plan == PRO
            else "dashboard"
        )

        code = "".join(secrets.choice(string.digits) for _ in range(6))
        user.verification_code = code
        user.verification_code_expires = datetime.utcnow() + timedelta(minutes=30)
        db.session.commit()

        with force_locale(user.preferred_language):
            email_sent = send_email(
                to=user.email,
                subject=gettext("Verifica tu correo en PATIA"),
                html=f"""
            <div style="font-family:Inter,Arial,sans-serif;max-width:600px;margin:0 auto;background:#0b1020;color:#eef3ff;padding:40px;border-radius:24px;">
                <img src="{_public_url('/static/img/logo-patia.png')}" style="width:160px;margin-bottom:24px;">
                <h1 style="color:#29d3a8;">{gettext("Verifica tu correo")}</h1>
                <p style="color:#9aa8c7;font-size:16px;">{gettext("Tu código de verificación es:")}</p>
                <div style="font-size:48px;font-weight:900;letter-spacing:12px;color:#fff;margin:24px 0;">{code}</div>
                <p style="color:#9aa8c7;font-size:14px;">{gettext("Este código expira en 30 minutos.")}</p>
            </div>
            """,
                language=user.preferred_language,
            )
        if not email_sent:
            flash(
                "No pudimos enviar el código. Usa Reenviar código para intentarlo nuevamente.",
                "danger",
            )

        return redirect(url_for("main.verify_email"))

    return render_template("auth.html", title=gettext("Crear cuenta"), button=gettext("Crear cuenta"), mode="register", plan=request.args.get("plan"), form_data={})


@main.route("/verify-email", methods=["GET", "POST"])
@limiter.limit("10 per 10 minutes", methods=["POST"])
def verify_email():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))

    if user.email_verified:
        return redirect(url_for("main.dashboard"))

    if request.method == "POST":
        code = request.form.get("code", "").strip()

        if not user.verification_code or not user.verification_code_expires:
            flash(gettext("Código inválido."), "danger")
            return redirect(url_for("main.verify_email"))

        if datetime.utcnow() > user.verification_code_expires:
            flash(gettext("El código expiró. Solicita uno nuevo."), "danger")
            return redirect(url_for("main.verify_email"))

        if code != user.verification_code:
            flash(gettext("Código incorrecto."), "danger")
            return redirect(url_for("main.verify_email"))

        user.email_verified = True
        user.verification_code = None
        user.verification_code_expires = None
        db.session.commit()
        with force_locale(user.preferred_language):
            send_email(
                to=user.email,
                subject=gettext("¡Bienvenido a PATIA!"),
                html=f"""
            <div style="font-family:Inter,Arial,sans-serif;max-width:600px;margin:0 auto;background:#0b1020;color:#eef3ff;padding:40px;border-radius:24px;">
                <img src="{_public_url('/static/img/logo-patia.png')}" style="width:160px;margin-bottom:24px;">
                <h1 style="color:#29d3a8;">{gettext("Bienvenido a PATIA, %(name)s!", name=user.first_name or user.company_name)}</h1>
                <p style="color:#9aa8c7;font-size:16px;line-height:1.6;">{gettext("Tu cuenta está lista. Tienes 14 días gratis para explorar todo.")}</p>
                <a href="{_public_url('/products')}" style="display:inline-block;margin-top:24px;padding:14px 28px;background:linear-gradient(135deg,#7c5cff,#29d3a8);color:white;text-decoration:none;border-radius:14px;font-weight:800;">{gettext("Ir a mi inventario")}</a>
            </div>
            """,
                language=user.preferred_language,
            )

        destination = session.pop("post_verify_destination", "dashboard")
        flash(gettext("¡Correo verificado! Bienvenido a PATIA."), "success")
        if destination == "subscribe":
            return redirect(url_for("main.subscribe"))
        return redirect(url_for("main.dashboard"))

    return render_template("verify_email.html", user=user)


@main.route("/resend-verification", methods=["POST"])
@limiter.limit("3 per 15 minutes")
def resend_verification():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))

    code = "".join(secrets.choice(string.digits) for _ in range(6))
    user.verification_code = code
    user.verification_code_expires = datetime.utcnow() + timedelta(minutes=30)
    db.session.commit()

    with force_locale(user.preferred_language):
        email_sent = send_email(
            to=user.email,
            subject=gettext("Nuevo código de verificación PATIA"),
            html=f"""
        <div style="font-family:Inter,Arial,sans-serif;max-width:600px;margin:0 auto;background:#0b1020;color:#eef3ff;padding:40px;border-radius:24px;">
            <h1 style="color:#29d3a8;">{gettext("Tu nuevo código")}</h1>
            <div style="font-size:48px;font-weight:900;letter-spacing:12px;color:#fff;margin:24px 0;">{code}</div>
            <p style="color:#9aa8c7;font-size:14px;">{gettext("Este código expira en 30 minutos.")}</p>
        </div>
        """,
            language=user.preferred_language,
        )
    if email_sent:
        flash(gettext("Te enviamos un nuevo código de verificación."), "success")
    else:
        flash(gettext("No pudimos enviar el código. Inténtalo nuevamente en unos minutos."), "danger")

    return redirect(url_for("main.verify_email"))


@main.route("/verification/change-email", methods=["POST"])
@limiter.limit("3 per 15 minutes")
def change_verification_email():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    if user.email_verified:
        return redirect(url_for("main.dashboard"))

    new_email = request.form.get("email", "").strip().lower()
    try:
        validate_email(new_email, check_deliverability=True)
    except EmailNotValidError:
        flash(gettext("El correo no es válido o no existe."), "danger")
        return redirect(url_for("main.verify_email"))

    duplicate = User.query.filter(
        func.lower(User.email) == new_email,
        User.id != user.id,
    ).first()
    if duplicate:
        flash(
            gettext("Ese correo ya pertenece a otra cuenta de PATIA."),
            "danger",
        )
        return redirect(url_for("main.verify_email"))

    code = "".join(secrets.choice(string.digits) for _ in range(6))
    user.email = new_email
    user.verification_code = code
    user.verification_code_expires = datetime.utcnow() + timedelta(minutes=30)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash(
            gettext("Ese correo ya pertenece a otra cuenta de PATIA."),
            "danger",
        )
        return redirect(url_for("main.verify_email"))

    with force_locale(user.preferred_language):
        email_sent = send_email(
            to=user.email,
            subject=gettext("Verifica tu correo en PATIA"),
            html=gettext(
                "Tu nuevo código de verificación es: %(code)s",
                code=code,
            ),
            language=user.preferred_language,
        )
    if email_sent:
        flash(
            gettext("Actualizamos tu correo y enviamos un código nuevo."),
            "success",
        )
    else:
        flash(
            gettext(
                "Actualizamos tu correo, pero no pudimos enviar el código. Intenta reenviarlo."
            ),
            "danger",
        )
    return redirect(url_for("main.verify_email"))


@main.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = User.query.filter_by(email=email).first()
        if user:
            token = secrets.token_urlsafe(36)
            user.reset_token = token
            user.reset_token_expires = datetime.utcnow() + timedelta(minutes=30)
            db.session.commit()
            with force_locale(user.preferred_language):
                email_sent = send_email(
                    to=user.email,
                    subject=gettext("Recupera tu contraseña PATIA"),
                    html=f"""
                <div style="font-family:Inter,Arial,sans-serif;max-width:600px;margin:0 auto;background:#0b1020;color:#eef3ff;padding:40px;border-radius:24px;">
                    <h1 style="color:#29d3a8;">{gettext("Recuperar contraseña")}</h1>
                    <p style="color:#9aa8c7;">{gettext("Haz clic en el botón para crear una nueva contraseña. Expira en 30 minutos.")}</p>
                    <a href="{_public_url(f'/reset-password/{token}')}" style="display:inline-block;margin-top:24px;padding:14px 28px;background:linear-gradient(135deg,#7c5cff,#29d3a8);color:white;text-decoration:none;border-radius:14px;font-weight:800;">{gettext("Crear nueva contraseña")}</a>
                </div>
                """,
                    language=user.preferred_language,
                )
            if not email_sent:
                flash("No pudimos enviar el enlace. Inténtalo nuevamente en unos minutos.", "danger")
                return redirect(url_for("main.forgot_password"))
        flash(gettext("Si ese correo existe, te enviamos un enlace."), "success")
        return redirect(url_for("main.login"))
    return render_template("forgot_password.html")


@main.route("/reset-password/<token>", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def reset_password(token):
    user = User.query.filter_by(reset_token=token).first()
    if (
        not user
        or not user.reset_token_expires
        or user.reset_token_expires < datetime.utcnow()
    ):
        return render_template("reset_password.html", token=None, expired=True)
    if request.method == "POST":
        password = request.form.get("password", "")
        if len(password) < 8:
            flash(gettext("La contraseña debe tener al menos 8 caracteres."), "danger")
            return render_template("reset_password.html", token=token, expired=False)
        user.set_password(password)
        user.reset_token = None
        user.reset_token_expires = None
        db.session.commit()
        flash(gettext("Contraseña actualizada. Inicia sesión."), "success")
        return redirect(url_for("main.login"))
    return render_template("reset_password.html", token=token, expired=False)


@main.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = User.query.filter_by(email=email).first()
        if not user or not user.check_password(password):
            flash(gettext("Correo o contraseña incorrectos."), "danger")
            return redirect(url_for("main.login"))

        token = secrets.token_hex(32)
        user.session_token = token
        membership = membership_for_login(user)
        if membership is None:
            db.session.rollback()
            flash(gettext("Tu acceso a esta empresa estÃ¡ desactivado. Contacta al propietario."), "danger")
            return redirect(url_for("main.login"))
        db.session.commit()
        session.clear()
        session["language"] = (
            user.preferred_language
            if user.preferred_language in SUPPORTED_LANGUAGES
            else "es"
        )
        session["user_id"] = user.id
        session["organization_id"] = membership.organization_id
        session["session_token"] = token
        access_user = membership.organization.owner
        days_used = (datetime.utcnow() - access_user.created_at).days
        if days_used >= 12 and not access_user.trial_warning_sent and not has_pro_access(access_user):
            with force_locale(access_user.preferred_language):
                warning_sent = send_email(
                    to=access_user.email,
                    subject=gettext("Tu prueba gratuita de PATIA termina en 2 días"),
                    html=f"""
                <div style="font-family:Inter,Arial,sans-serif;max-width:600px;margin:0 auto;background:#0b1020;color:#eef3ff;padding:40px;border-radius:24px;">
                    <img src="{_public_url('/static/img/logo-patia.png')}" style="width:160px;margin-bottom:24px;">
                    <h1 style="color:#ff5c7a;">{gettext("Tu prueba termina en 2 días")}</h1>
                    <p style="color:#9aa8c7;font-size:16px;line-height:1.6;">{gettext("Hola %(name)s, tu periodo de prueba gratuita de PATIA termina pronto. No pierdas el acceso a tu inventario y ventas.", name=access_user.first_name or access_user.company_name)}</p>
                    <a href="{_public_url('/subscribe')}" style="display:inline-block;margin-top:24px;padding:14px 28px;background:linear-gradient(135deg,#7c5cff,#29d3a8);color:white;text-decoration:none;border-radius:14px;font-weight:800;">{gettext("Activar PATIA Pro")}</a>
                </div>
                """,
                    language=access_user.preferred_language,
                )
            if warning_sent:
                access_user.trial_warning_sent = True
                db.session.commit()
        flash(gettext("Sesión iniciada correctamente."), "success")
        return redirect(_safe_next_url(request.args.get("next")))

    return render_template("auth.html", title=gettext("Iniciar sesión"), button=gettext("Entrar"), mode="login")


@main.get("/logout")
def logout_get():
    if current_user():
        flash(
            gettext(
                "Para cerrar sesión de forma segura, utiliza el botón "
                "Cerrar sesión."
            ),
            "info",
        )
        return redirect(url_for("main.dashboard"))
    flash(
        gettext("Tu sesión ya terminó. Inicia sesión nuevamente."),
        "info",
    )
    return redirect(url_for("main.login"))


@main.post("/logout")
def logout():
    language = session.get("language", "es")
    session.clear()
    session["language"] = language if language in SUPPORTED_LANGUAGES else "es"
    flash(gettext("Has cerrado sesión correctamente."), "success")
    return redirect(url_for("main.login"))


@main.app_template_filter("money")
def money_filter(value, currency_code=None, locale_code=None):
    return money(value or 0, currency_code, locale_code)


@main.app_template_filter("compact_money")
def compact_money_filter(value):
    amount = money_decimal(value or 0, nonnegative=False)
    return money(amount)


def _dashboard_sales_summary(
    organization_id, start_at, end_before, currency_code
):
    """Return sales and grouped-ticket totals for one dashboard period."""
    totals = (
        db.session.query(
            func.coalesce(func.sum(Sale.total), 0).label("sales"),
            (
                func.count(func.distinct(Sale.ticket_id))
                + func.coalesce(
                    func.sum(
                        case((Sale.ticket_id.is_(None), 1), else_=0)
                    ),
                    0,
                )
            ).label("tickets"),
        )
        .filter(
            Sale.organization_id == organization_id,
            Sale.currency_code == currency_code,
            Sale.created_at >= start_at,
            Sale.created_at < end_before,
        )
        .one()
    )
    sales = money_decimal(totals.sales or 0)
    tickets = int(totals.tickets or 0)
    return {
        "sales": sales,
        "tickets": tickets,
        "average_ticket": money_decimal(
            sales / tickets if tickets else MONEY_ZERO
        ),
    }


def _dashboard_percentage_change(current, previous):
    previous = money_decimal(previous or 0, nonnegative=False)
    if previous == MONEY_ZERO:
        return None
    current = money_decimal(current or 0, nonnegative=False)
    return round((current - previous) / previous * 100, 1)


def analytics(user=None):
    user = user or current_user()
    organization_id = current_organization_id(user)
    cash_session = open_cash_session(organization_id)
    membership = active_membership(user)
    organization = membership.organization
    currency_code, _ = organization_money_context(organization)
    timezone_name = safe_timezone_name(
        membership.organization.timezone if membership else DEFAULT_TIMEZONE
    )
    today = local_today(timezone_name)
    start, tomorrow = local_date_bounds_utc(
        today,
        today + timedelta(days=1),
        timezone_name,
    )
    week_start, week_end = local_date_bounds_utc(
        today - timedelta(days=6),
        today + timedelta(days=1),
        timezone_name,
    )
    month_start_date = today.replace(day=1)
    previous_month_end_date = month_start_date
    previous_month_last_date = month_start_date - timedelta(days=1)
    previous_month_start_date = previous_month_last_date.replace(day=1)
    month_start, month_end = local_date_bounds_utc(
        month_start_date,
        today + timedelta(days=1),
        timezone_name,
    )
    previous_month_start, previous_month_end = local_date_bounds_utc(
        previous_month_start_date,
        previous_month_end_date,
        timezone_name,
    )

    products = Product.query.filter_by(organization_id=organization_id, is_active=True).all()

    total_products = len(products)
    total_sales = Sale.query.filter_by(organization_id=organization_id).count()
    inventory_value = (
        db.session.query(func.sum(Product.stock * Product.cost_price))
        .filter(Product.organization_id == organization_id, Product.is_active.is_(True))
        .scalar()
        or 0
    )
    low_stock_products = sorted(
        (p for p in products if p.stock <= p.min_stock),
        key=lambda p: (p.stock - p.min_stock, p.name.casefold()),
    )
    low_stock = len(low_stock_products)
    out_of_stock = sum(1 for product in products if product.stock <= 0)

    latest_credit_ids = (
        db.session.query(
            CustomerCreditMovement.customer_id,
            func.max(CustomerCreditMovement.id).label("last_id"),
        )
        .filter(CustomerCreditMovement.organization_id == organization_id)
        .group_by(CustomerCreditMovement.customer_id)
        .subquery()
    )
    total_credit_pending = (
        db.session.query(func.sum(CustomerCreditMovement.balance_after))
        .join(
            latest_credit_ids,
            CustomerCreditMovement.id == latest_credit_ids.c.last_id,
        )
        .scalar()
        or MONEY_ZERO
    )

    today_summary = _dashboard_sales_summary(
        organization_id,
        start,
        tomorrow,
        currency_code,
    )
    month_summary = _dashboard_sales_summary(
        organization_id,
        month_start,
        month_end,
        currency_code,
    )
    previous_month_summary = _dashboard_sales_summary(
        organization_id,
        previous_month_start,
        previous_month_end,
        currency_code,
    )
    dashboard_summary = {
        "out_of_stock": out_of_stock,
        "today_sales": today_summary["sales"],
        "today_tickets": today_summary["tickets"],
        "month_sales": month_summary["sales"],
        "month_tickets": month_summary["tickets"],
        "month_sales_change": _dashboard_percentage_change(
            month_summary["sales"],
            previous_month_summary["sales"],
        ),
        "month_average_ticket": month_summary["average_ticket"],
        "month_average_change": _dashboard_percentage_change(
            month_summary["average_ticket"],
            previous_month_summary["average_ticket"],
        ),
    }

    week_sales = db.session.query(func.sum(Sale.total)).filter(
        Sale.organization_id == organization_id,
        Sale.currency_code == currency_code,
        Sale.created_at >= week_start,
        Sale.created_at < week_end,
    ).scalar() or 0

    profit = (
        db.session.query(
            func.sum(
                case(
                    (
                        Sale.unit_cost.is_not(None),
                        (Sale.unit_price - Sale.unit_cost) * Sale.quantity,
                    ),
                    else_=0,
                )
            )
        )
        .filter(
            Sale.organization_id == organization_id,
            Sale.currency_code == currency_code,
            Sale.created_at >= week_start,
            Sale.created_at < week_end,
        )
        .scalar() or 0
    )

    top_products = (
        db.session.query(
            Product.id,
            Product.name,
            func.sum(Sale.quantity).label("qty"),
            func.sum(Sale.total).label("revenue"),
        )
        .join(Sale)
        .filter(
            Product.organization_id == organization_id,
            Sale.currency_code == currency_code,
        )
        .group_by(Product.id)
        .order_by(func.sum(Sale.quantity).desc())
        .limit(5)
        .all()
    )

    category_sales = (
        db.session.query(Product.category, func.sum(Sale.total).label("revenue"))
        .join(Sale)
        .filter(
            Product.organization_id == organization_id,
            Sale.currency_code == currency_code,
        )
        .group_by(Product.category)
        .order_by(func.sum(Sale.total).desc())
        .all()
    )

    sold_by_product = dict(
        db.session.query(Sale.product_id, func.sum(Sale.quantity))
        .filter(
            Sale.organization_id == organization_id,
            Sale.created_at >= week_start,
            Sale.created_at < week_end,
        )
        .group_by(Sale.product_id)
        .all()
    )

    alerts = []
    for p in products:
        sold_7_days = sold_by_product.get(p.id, 0) or 0
        avg_daily_sales = sold_7_days / 7
        days_left = round(p.stock / avg_daily_sales, 1) if avg_daily_sales > 0 else None

        if p.stock <= 0:
            alerts.append({
                "type": "critical",
                "title": gettext("%(product)s agotado", product=p.name),
                "text": gettext("Stock actual: 0. Necesitas reabastecerlo inmediatamente."),
            })
        elif p.stock <= p.min_stock:
            alerts.append({
                "type": "critical",
                "title": gettext("Reordenar %(product)s", product=p.name),
                "text": gettext(
                    "Stock actual: %(stock)s. Mínimo recomendado: %(minimum)s.",
                    stock=p.stock,
                    minimum=p.min_stock,
                ),
            })
        elif days_left is not None and days_left <= 3:
            alerts.append({
                "type": "critical",
                "title": gettext("%(product)s se agotará pronto", product=p.name),
                "text": gettext(
                    "Con el ritmo actual de ventas, se acabará en aproximadamente %(days)s días.",
                    days=days_left,
                ),
            })
        elif days_left is not None and days_left <= 7:
            alerts.append({
                "type": "warning",
                "title": gettext("Vigilar %(product)s", product=p.name),
                "text": gettext("Inventario estimado para %(days)s días.", days=days_left),
            })

    recommendations = []
    if top_products:
        recommendations.append({
            "text": gettext(
                "%(product)s es el producto con más movimiento. Conviene revisar sus existencias antes de tu próxima compra.",
                product=top_products[0].name,
            ),
            "source": gettext("Ventas registradas en los últimos 7 días."),
            "action": gettext(
                "Confirma sus existencias, costo y stock mínimo antes de preparar tu siguiente compra."
            ),
            "label": gettext("Abrir producto"),
            "url": url_for(
                "main.edit_product", product_id=top_products[0].id
            ),
        })
    if week_sales > 0:
        recommendations.append({
            "text": gettext(
                "En los últimos 7 días registraste %(amount)s en ventas.",
                amount=money(week_sales),
            ),
            "source": gettext("Suma de las ventas registradas en el periodo."),
            "action": gettext(
                "Compara los días del periodo para identificar cuándo vendes más."
            ),
            "label": gettext("Ver reportes"),
            "url": url_for("main.reports", period="7d"),
        })
    if profit > 0:
        recommendations.append({
            "text": gettext(
                "Tu utilidad estimada de los últimos 7 días es %(amount)s, considerando el costo registrado de los productos vendidos.",
                amount=money(profit),
            ),
            "source": gettext("Precio vendido menos costo histórico conocido."),
            "action": gettext(
                "Revisa qué productos aportaron esa utilidad y cuáles necesitan un mejor margen."
            ),
            "label": gettext("Revisar utilidad"),
            "url": url_for("main.reports", period="7d"),
        })
    if low_stock:
        recommendations.append({
            "text": (
                gettext(
                    "Tienes %(count)s producto con inventario bajo. Revísalo antes de que afecte una venta.",
                    count=low_stock,
                )
                if low_stock == 1
                else gettext(
                    "Tienes %(count)s productos con inventario bajo. Revísalos antes de que afecten una venta.",
                    count=low_stock,
                )
            ),
            "source": gettext("Stock actual comparado con el mínimo de cada producto."),
            "action": gettext(
                "Reabastece primero los productos agotados o más cercanos a quedarse sin existencias."
            ),
            "label": gettext("Ver qué surtir"),
            "url": url_for("main.products", low_stock=1),
        })

    alerts = alerts[:5]
    return dict(
        total_products=total_products,
        total_sales=total_sales,
        inventory_value=inventory_value,
        low_stock=low_stock,
        low_stock_products=low_stock_products,
        today_sales=today_summary["sales"],
        dashboard_summary=dashboard_summary,
        week_sales=week_sales,
        profit=profit,
        cash_session=cash_session,
        cash_expected=(
            cash_expected_amount(cash_session.id) if cash_session else None
        ),
        total_credit_pending=total_credit_pending,
        top_products=top_products,
        category_sales=category_sales,
        alerts=alerts,
        recommendations=recommendations,
    )


REPORT_PERIODS = {
    "today",
    "7d",
    "30d",
    "this_month",
    "previous_month",
    "custom",
}


def _parse_report_period(
    args,
    *,
    today=None,
    timezone_name=DEFAULT_TIMEZONE,
    now_utc=None,
):
    """Return a validated, half-open date interval for report queries."""
    timezone_name = safe_timezone_name(timezone_name)
    today = today or local_today(timezone_name, now_utc=now_utc)
    period = (args.get("period") or "7d").strip()
    error = None
    custom_start = (args.get("start") or "").strip()
    custom_end = (args.get("end") or "").strip()

    if period not in REPORT_PERIODS:
        period = "7d"
        error = gettext("El periodo solicitado no es válido. Mostramos los últimos 7 días.")

    if period == "today":
        start_date = end_date = today
    elif period == "30d":
        start_date, end_date = today - timedelta(days=29), today
    elif period == "this_month":
        start_date, end_date = today.replace(day=1), today
    elif period == "previous_month":
        current_month = today.replace(day=1)
        end_date = current_month - timedelta(days=1)
        start_date = end_date.replace(day=1)
    elif period == "custom":
        try:
            start_date = datetime.strptime(custom_start, "%Y-%m-%d").date()
            end_date = datetime.strptime(custom_end, "%Y-%m-%d").date()
            if start_date > end_date:
                raise ValueError("start_after_end")
            if (end_date - start_date).days > 365:
                raise ValueError("range_too_long")
        except (TypeError, ValueError) as exc:
            period = "7d"
            start_date, end_date = today - timedelta(days=6), today
            if str(exc) == "start_after_end":
                error = gettext("La fecha inicial no puede ser posterior a la fecha final.")
            elif str(exc) == "range_too_long":
                error = gettext("El rango personalizado no puede superar 366 días.")
            else:
                error = gettext("Escribe fechas válidas. Mostramos los últimos 7 días.")
    else:
        start_date, end_date = today - timedelta(days=6), today

    start_at, end_before = local_date_bounds_utc(
        start_date,
        end_date + timedelta(days=1),
        timezone_name,
    )
    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "start_at": start_at,
        "end_before": end_before,
        "custom_start": custom_start,
        "custom_end": custom_end,
        "error": error,
    }


def _report_analytics(
    organization_id,
    period,
    *,
    timezone_name=DEFAULT_TIMEZONE,
    currency_code=None,
    include_mixed_currency=False,
):
    timezone_name = safe_timezone_name(timezone_name)
    start_at = period["start_at"]
    end_before = period["end_before"]
    sale_filters = [
        Sale.organization_id == organization_id,
        Sale.created_at >= start_at,
        Sale.created_at < end_before,
    ]
    if currency_code:
        sale_filters.append(Sale.currency_code == currency_code)
    sale_filters = tuple(sale_filters)
    known_cost = Sale.unit_cost.is_not(None)
    line_profit = Sale.total - (Sale.unit_cost * Sale.quantity)

    totals = (
        db.session.query(
            func.coalesce(func.sum(Sale.total), 0).label("sales"),
            func.coalesce(
                func.sum(case((known_cost, line_profit), else_=0)),
                0,
            ).label("profit"),
            func.coalesce(
                func.sum(case((known_cost, Sale.total), else_=0)),
                0,
            ).label("known_revenue"),
            func.coalesce(
                func.sum(case((Sale.unit_cost.is_(None), 1), else_=0)),
                0,
            ).label("unknown_cost_lines"),
        )
        .filter(*sale_filters)
        .one()
    )
    ticket_count = (
        db.session.query(
            func.count(func.distinct(Sale.ticket_id))
            + func.coalesce(
                func.sum(case((Sale.ticket_id.is_(None), 1), else_=0)),
                0,
            )
        )
        .filter(*sale_filters)
        .scalar()
        or 0
    )
    total_sales = money_decimal(totals.sales or 0)
    known_profit = money_decimal(totals.profit or 0, nonnegative=False)
    known_revenue = money_decimal(totals.known_revenue or 0)

    if db.session.get_bind().dialect.name == "postgresql":
        hour_value = func.date_trunc("hour", Sale.created_at)
    else:
        hour_value = func.strftime(
            "%Y-%m-%d %H:00:00",
            Sale.created_at,
        )
    daily_rows = (
        db.session.query(
            hour_value.label("hour"),
            func.sum(Sale.total).label("sales"),
            func.sum(case((known_cost, line_profit), else_=0)).label("profit"),
        )
        .filter(*sale_filters)
        .group_by(hour_value)
        .order_by(hour_value)
        .all()
    )
    daily_lookup = {}
    for row in daily_rows:
        hour = row.hour
        if isinstance(hour, str):
            hour = datetime.fromisoformat(hour)
        day_key = utc_to_local(hour, timezone_name).date().isoformat()
        sales_value, profit_value = daily_lookup.get(
            day_key, (MONEY_ZERO, MONEY_ZERO)
        )
        daily_lookup[day_key] = (
            sales_value + money_decimal(row.sales or 0),
            profit_value + money_decimal(row.profit or 0, nonnegative=False),
        )
    daily = []
    cursor = period["start_date"]
    while cursor <= period["end_date"]:
        sales_value, profit_value = daily_lookup.get(
            cursor.isoformat(), (MONEY_ZERO, MONEY_ZERO)
        )
        daily.append({
            "date": cursor.isoformat(),
            "sales": sales_value,
            "profit": profit_value,
        })
        cursor += timedelta(days=1)

    payment_key = func.coalesce(
        SalesTicket.payment_method,
        Sale.payment_method,
        "other",
    )
    payment_rows = (
        db.session.query(
            payment_key.label("method"),
            func.sum(Sale.total).label("amount"),
            (
                func.count(func.distinct(Sale.ticket_id))
                + func.coalesce(
                    func.sum(case((Sale.ticket_id.is_(None), 1), else_=0)),
                    0,
                )
            ).label("tickets"),
        )
        .select_from(Sale)
        .outerjoin(SalesTicket, Sale.sales_ticket_id == SalesTicket.id)
        .filter(*sale_filters)
        .group_by(payment_key)
        .all()
    )
    payment_by_key = {
        row.method: {
            "amount": money_decimal(row.amount or 0),
            "tickets": int(row.tickets or 0),
        }
        for row in payment_rows
    }
    payments = []
    for method in PAYMENT_METHOD_LABELS:
        item = payment_by_key.get(
            method, {"amount": MONEY_ZERO, "tickets": 0}
        )
        payments.append({
            "key": method,
            "label": _payment_method_label(method),
            "amount": item["amount"],
            "tickets": item["tickets"],
            "percentage": round(
                item["amount"] / total_sales * 100,
                1,
            ) if total_sales else 0,
        })

    top_selling = (
        db.session.query(
            Product.name.label("name"),
            func.sum(Sale.quantity).label("units"),
            func.sum(Sale.total).label("revenue"),
        )
        .join(Sale, Sale.product_id == Product.id)
        .filter(*sale_filters, Product.organization_id == organization_id)
        .group_by(Product.id, Product.name)
        .order_by(func.sum(Sale.quantity).desc(), Product.name)
        .limit(10)
        .all()
    )

    unknown_lines = func.sum(
        case((Sale.unit_cost.is_(None), 1), else_=0)
    )
    product_profit = func.sum(
        case((known_cost, line_profit), else_=0)
    )
    profitable_rows = (
        db.session.query(
            Product.name.label("name"),
            func.sum(Sale.quantity).label("units"),
            func.sum(Sale.total).label("revenue"),
            func.sum(
                case(
                    (known_cost, Sale.unit_cost * Sale.quantity),
                    else_=0,
                )
            ).label("cost"),
            product_profit.label("profit"),
            unknown_lines.label("unknown_lines"),
        )
        .join(Sale, Sale.product_id == Product.id)
        .filter(*sale_filters, Product.organization_id == organization_id)
        .group_by(Product.id, Product.name)
        .order_by(unknown_lines, product_profit.desc(), Product.name)
        .limit(10)
        .all()
    )
    profitable_products = []
    for row in profitable_rows:
        revenue = money_decimal(row.revenue or 0)
        has_unknown_cost = bool(row.unknown_lines)
        profit = (
            None
            if has_unknown_cost
            else money_decimal(row.profit or 0, nonnegative=False)
        )
        profitable_products.append({
            "name": row.name,
            "units": int(row.units or 0),
            "revenue": revenue,
            "cost": (
                None
                if has_unknown_cost
                else money_decimal(row.cost or 0)
            ),
            "profit": profit,
            "margin": (
                round(profit / revenue * 100, 1)
                if profit is not None and revenue
                else None
            ),
        })

    return {
        "currency_code": currency_code,
        "mixed_currency_lines": (
            Sale.query.filter(
                Sale.organization_id == organization_id,
                Sale.created_at >= start_at,
                Sale.created_at < end_before,
                Sale.currency_code != currency_code,
            ).count()
            if currency_code and include_mixed_currency else 0
        ),
        "report_period": period,
        "report_kpis": {
            "sales": total_sales,
            "profit": known_profit,
            "margin": (
                round(known_profit / known_revenue * 100, 1)
                if known_revenue
                else None
            ),
            "average_ticket": money_decimal(
                total_sales / ticket_count if ticket_count else MONEY_ZERO
            ),
            "ticket_count": int(ticket_count),
            "profit_coverage": (
                round(known_revenue / total_sales * 100, 1)
                if total_sales
                else None
            ),
        },
        "unknown_cost_lines": int(totals.unknown_cost_lines or 0),
        "daily_report": daily,
        "payments_report": payments,
        "top_selling_report": top_selling,
        "profitable_products_report": profitable_products,
    }


@main.route("/")
def dashboard():
    user = current_user()
    if not user:
        if any(
            session.get(flag)
            for flag in (
                "kicked_out",
                "membership_disabled",
                "session_expired",
            )
        ):
            return authentication_required_response()
        language = session.get("language", "es")
        session.clear()
        session["language"] = language if language in SUPPORTED_LANGUAGES else "es"
        from .plans import commercial_plans

        return render_template(
            "landing.html",
            commercial_plans=commercial_plans(current_app.config),
        )
    membership = active_membership(user)
    if not has_permission(membership, "view_dashboard"):
        return redirect(url_for("main.sell"))
    access_block = _trial_access_response(user)
    if access_block:
        return access_block
    dashboard_data = analytics(user)
    owner = membership.organization.owner
    from .plans import has_entitlement
    from .pro.services import build_executive_dashboard

    can_use_advanced_reports = has_entitlement(owner, "advanced_reports")
    dashboard_executive = build_executive_dashboard(
        membership.organization,
        {"period": "7d"},
    )
    has_basic_data = all(
        (owner.company_name, owner.phone, owner.city, owner.state)
    )
    has_products = dashboard_data["total_products"] > 0
    has_sales = dashboard_data["total_sales"] > 0
    onboarding_steps = [
        {
            "title": gettext("Completa los datos básicos"),
            "text": gettext("Confirma la información principal de tu negocio."),
            "completed": has_basic_data,
            "action_label": (
                gettext("Completar datos")
                if membership.role == "OWNER"
                else None
            ),
            "action_url": (
                url_for("main.settings")
                if membership.role == "OWNER"
                else None
            ),
        },
        {
            "title": gettext("Agrega tu primer producto"),
            "text": gettext("Captura un producto o importa tu catálogo desde Excel."),
            "completed": has_products,
            "action_label": gettext("Agregar primer producto"),
            "action_url": url_for("main.products"),
        },
        {
            "title": gettext("Registra tu primera venta"),
            "text": gettext("Prueba el punto de venta con un producto de tu catálogo."),
            "completed": has_sales,
            "action_label": gettext("Registrar primera venta") if has_products else None,
            "action_url": url_for("main.sell") if has_products else None,
        },
        {
            "title": gettext("Revisa tus resultados"),
            "text": gettext("Consulta ventas, inventario y alertas en este panel."),
            "completed": has_sales,
            "action_label": None,
            "action_url": None,
        },
    ]
    completed_steps = sum(step["completed"] for step in onboarding_steps)
    onboarding_progress = round(completed_steps / len(onboarding_steps) * 100)
    onboarding_completed = completed_steps == len(onboarding_steps)

    return render_template(
        "dashboard.html",
        company_name=membership.organization.name,
        user=user,
        trial_days_left=max(0, 14 - (datetime.utcnow() - owner.created_at).days) if owner.created_at else 14,
        onboarding_steps=onboarding_steps,
        onboarding_completed=onboarding_completed,
        onboarding_progress=onboarding_progress,
        show_onboarding=not has_products or not has_sales,
        dashboard_executive=dashboard_executive,
        can_use_advanced_reports=can_use_advanced_reports,
        can_edit_goal=has_permission(membership, "manage_subscription"),
        **dashboard_data,
    )


@main.route("/products")
@require_permission("manage_inventory")
def products():
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    if trial_expired(user):
        return render_template("trial_expired.html")
    organization_id = current_organization_id(user)
    q = request.args.get("q", "").strip()
    low_stock_only = request.args.get("low_stock") == "1"
    in_stock_only = request.args.get("in_stock") == "1"
    out_of_stock_only = request.args.get("out_of_stock") == "1"
    no_sales_only = request.args.get("no_sales") == "1"
    missing_cost_only = request.args.get("missing_cost") == "1"
    low_margin_only = request.args.get("low_margin") == "1"
    catalog_query = Product.query.filter(
        Product.organization_id == organization_id,
        Product.is_active.is_(True),
    )
    catalog_count = catalog_query.count()
    low_stock_count = catalog_query.filter(
        Product.stock <= Product.min_stock,
    ).count()
    query = catalog_query
    if low_stock_only:
        query = query.filter(Product.stock <= Product.min_stock)
        if in_stock_only:
            query = query.filter(Product.stock > 0)
    if out_of_stock_only:
        query = query.filter(Product.stock <= 0)
    if no_sales_only:
        timezone_name = safe_timezone_name(
            active_membership(user).organization.timezone
        )
        today = local_today(timezone_name)
        try:
            start_date = datetime.strptime(
                request.args.get("start", ""), "%Y-%m-%d"
            ).date()
            end_date = datetime.strptime(
                request.args.get("end", ""), "%Y-%m-%d"
            ).date()
        except (TypeError, ValueError):
            start_date = today - timedelta(days=29)
            end_date = today
        start_at, end_before = local_date_bounds_utc(
            start_date,
            end_date + timedelta(days=1),
            timezone_name,
        )
        sold_in_period = exists().where(
            Sale.organization_id == organization_id,
            Sale.product_id == Product.id,
            Sale.created_at >= start_at,
            Sale.created_at < end_before,
        )
        query = query.filter(~sold_in_period)
    if missing_cost_only:
        query = query.filter(
            (Product.cost_price.is_(None)) | (Product.cost_price <= 0)
        )
    if low_margin_only:
        query = query.filter(
            Product.sale_price > 0,
            Product.cost_price.is_not(None),
            ((Product.sale_price - Product.cost_price) / Product.sale_price)
            <= Decimal("0.15"),
        )
    if q:
        query = query.filter(
            Product.name.ilike(f"%{q}%") |
            Product.category.ilike(f"%{q}%") |
            Product.sku.ilike(f"%{q}%")
        )
    result_count = query.count()
    per_page = 100
    page_count = max(1, (result_count + per_page - 1) // per_page)
    requested_page = request.args.get("page", 1, type=int) or 1
    page = min(max(requested_page, 1), page_count)
    products_result = (
        query.order_by(Product.name)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    page_start = ((page - 1) * per_page + 1) if result_count else 0
    page_end = min(page * per_page, result_count)
    active_filter_label = None
    filter_source = request.args.get("source")
    if out_of_stock_only:
        active_filter_label = gettext("Productos agotados")
    elif low_stock_only:
        active_filter_label = gettext("Productos por agotarse")
    elif no_sales_only:
        active_filter_label = gettext("Productos sin ventas en el periodo")
    elif missing_cost_only:
        active_filter_label = gettext("Productos sin costo conocido")
    elif low_margin_only:
        active_filter_label = gettext("Productos con margen de 15%% o menos")
    return render_template(
        "products.html",
        products=products_result,
        catalog_count=catalog_count,
        result_count=result_count,
        page=page,
        page_count=page_count,
        page_start=page_start,
        page_end=page_end,
        low_stock_count=low_stock_count,
        low_stock_only=low_stock_only,
        in_stock_only=in_stock_only,
        out_of_stock_only=out_of_stock_only,
        no_sales_only=no_sales_only,
        missing_cost_only=missing_cost_only,
        low_margin_only=low_margin_only,
        active_filter_label=active_filter_label,
        filter_source=filter_source,
        q=q,
        user=user,
    )


def _quick_load_product_payload(product):
    return {
        "id": product.id,
        "barcode": product.barcode,
        "name": product.name,
        "sku": product.sku,
        "stock": product.stock,
        "min_stock": product.min_stock,
        "sale_price": money_json(product.sale_price),
        "sale_price_display": money(product.sale_price),
        "supplier": product.supplier,
        "is_active": product.is_active,
        "edit_url": (
            url_for("main.edit_product", product_id=product.id)
            if product.is_active
            else None
        ),
    }


@main.route("/products/quick-load")
@require_permission("manage_inventory")
def quick_load_products():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    access_block = _trial_access_response(user)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    suppliers = {
        name
        for (name,) in db.session.query(Supplier.name)
        .filter(Supplier.organization_id == organization_id)
        .all()
        if name
    }
    suppliers.update(
        name
        for (name,) in db.session.query(Product.supplier)
        .filter(
            Product.organization_id == organization_id,
            Product.supplier.isnot(None),
        )
        .distinct()
        .all()
        if name
    )
    categories = [
        name
        for (name,) in db.session.query(Product.category)
        .filter(
            Product.organization_id == organization_id,
            Product.category.isnot(None),
        )
        .distinct()
        .order_by(Product.category)
        .all()
        if name
    ]
    return render_template(
        "quick_load.html",
        user=user,
        suppliers=sorted(suppliers, key=str.casefold),
        categories=categories,
    )


@main.route("/api/products/quick-load/lookup")
@require_permission("manage_inventory")
def quick_load_lookup():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": gettext("Inicia sesión para continuar.")}), 401
    access_block = _trial_access_response(user, json_response=True)
    if access_block:
        return access_block

    try:
        barcode = normalize_barcode(request.args.get("barcode"))
    except ValueError:
        return jsonify({
            "ok": False,
            "error": gettext("Escanea o escribe un código de barras válido."),
        }), 400

    organization_id = current_organization_id(user)
    product = find_company_product_by_barcode(organization_id, barcode)
    if product:
        return jsonify({
            "ok": True,
            "found": True,
            "product": _quick_load_product_payload(product),
        })

    metadata = lookup_barcode(barcode)
    return jsonify({
        "ok": True,
        "found": False,
        "barcode": barcode,
        "suggested_sku": automatic_sku(organization_id, barcode),
        "metadata": (
            {"name": metadata.name, "category": metadata.category}
            if metadata
            else None
        ),
    })


@main.route("/api/products/quick-load", methods=["POST"])
@require_permission("manage_inventory")
def quick_load_create_product():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": gettext("Inicia sesión para continuar.")}), 401
    access_block = _trial_access_response(user, json_response=True)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    membership = active_membership(user)
    owner = current_organization_owner(user)
    payload = request.get_json(silent=True) or {}
    try:
        barcode = normalize_barcode(payload.get("barcode"))
        name = str(payload.get("name") or "").strip()
        sku = str(payload.get("sku") or "").strip() or automatic_sku(organization_id, barcode)
        category = str(payload.get("category") or "").strip() or "General"
        supplier = str(payload.get("supplier") or "").strip() or None
        cost_price = money_decimal(payload.get("cost_price") or 0)
        sale_price = money_decimal(payload.get("sale_price") or 0)
        stock = int(payload.get("stock") or 0)
        min_stock = int(payload.get("min_stock") or 0)
    except (TypeError, ValueError, OverflowError):
        return jsonify({
            "ok": False,
            "error": gettext("Revisa los datos del producto e inténtalo nuevamente."),
        }), 400

    if (
        not name
        or not sku
        or len(name) > 160
        or len(sku) > 64
        or len(category) > 80
        or (supplier and len(supplier) > 120)
        or stock < 0
        or min_stock < 0
    ):
        return jsonify({
            "ok": False,
            "error": gettext("Revisa los datos del producto e inténtalo nuevamente."),
        }), 400

    existing = find_company_product_by_barcode(organization_id, barcode)
    if existing:
        return jsonify({
            "ok": False,
            "duplicate": True,
            "error": gettext("Ese código ya pertenece a un producto de tu inventario."),
            "product": _quick_load_product_payload(existing),
        }), 409
    if Product.query.filter_by(organization_id=organization_id, sku=sku).first():
        return jsonify({
            "ok": False,
            "error": gettext("Ese SKU ya está en uso. Escribe uno diferente."),
        }), 409

    product = Product(
        organization_id=organization_id,
        user_id=owner.id,
        barcode=barcode,
        sku=sku,
        name=name,
        category=category,
        supplier=supplier,
        cost_price=cost_price,
        sale_price=sale_price,
        stock=stock,
        min_stock=min_stock,
    )
    db.session.add(product)
    try:
        record_opening_balance(
            product,
            membership,
            reason=gettext("Alta mediante carga rápida"),
        )
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        existing = find_company_product_by_barcode(organization_id, barcode)
        current_app.logger.info(
            "Carga rápida rechazada por identificador duplicado para organization_id=%s",
            organization_id,
        )
        response = {
            "ok": False,
            "duplicate": bool(existing),
            "error": gettext(
                "El código o el SKU ya fue registrado. Escanéalo de nuevo para continuar."
            ),
        }
        if existing:
            response["product"] = _quick_load_product_payload(existing)
        return jsonify(response), 409

    return jsonify({
        "ok": True,
        "message": gettext("%(product)s se agregó al inventario.", product=product.name),
        "product": _quick_load_product_payload(product),
    }), 201


@main.route("/api/products/<int:product_id>/quick-restock", methods=["POST"])
@require_permission("make_inventory_adjustments")
def quick_load_restock_product(product_id):
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": gettext("Inicia sesión para continuar.")}), 401
    access_block = _trial_access_response(user, json_response=True)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    membership = active_membership(user)
    owner = current_organization_owner(user)
    payload = request.get_json(silent=True) or {}
    try:
        quantity = int(payload.get("quantity"))
    except (TypeError, ValueError, OverflowError):
        quantity = 0
    if quantity <= 0:
        return jsonify({
            "ok": False,
            "error": gettext("La cantidad recibida debe ser mayor que cero."),
        }), 400

    product = (
        Product.query.filter_by(id=product_id, organization_id=organization_id)
        .with_for_update()
        .first()
    )
    if not product:
        return jsonify({
            "ok": False,
            "error": gettext("No encontramos ese producto en tu inventario."),
        }), 404

    try:
        if product.stock > 2_147_483_647 - quantity:
            return jsonify({
                "ok": False,
                "error": gettext("La cantidad recibida supera el límite permitido."),
            }), 400
        stock_before = product.stock
        product.stock = stock_before + quantity
        product.is_active = True
        restock_event = InventoryRestockEvent(
            organization_id=organization_id,
            user_id=owner.id,
            product_id=product.id,
            quantity=quantity,
            stock_before=stock_before,
            stock_after=product.stock,
        )
        db.session.add(restock_event)
        db.session.flush()
        record_inventory_movement(
            product,
            membership,
            "RESTOCK",
            stock_before,
            product.stock,
            reason=gettext("Reabastecimiento desde carga rápida"),
            restock_event=restock_event,
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "No se pudo reabastecer desde carga rápida para organization_id=%s product_id=%s",
            organization_id,
            product_id,
        )
        return jsonify({
            "ok": False,
            "error": gettext("No pudimos actualizar el inventario. Inténtalo nuevamente."),
        }), 500

    return jsonify({
        "ok": True,
        "message": gettext(
            "Se agregaron %(quantity)s unidades a %(product)s.",
            quantity=quantity,
            product=product.name,
        ),
        "product": _quick_load_product_payload(product),
    })


@main.route("/download-template")
@require_permission("manage_inventory")
def download_template():
    if not session.get("user_id"):
        return redirect(url_for("main.login"))

    import pandas as pd

    columns = [
        gettext("SKU"),
        gettext("Código de barras"),
        gettext("Nombre del producto"),
        gettext("Categoría"),
        gettext("Proveedor"),
        gettext("Costo"),
        gettext("Precio de venta"),
        gettext("Stock inicial"),
        gettext("Stock mínimo"),
    ]
    df = pd.DataFrame(columns=columns)
    example_rows = [
        [
            "FER-0001",
            "07501234560001",
            gettext("Taladro percutor 1/2 pulgada"),
            gettext("Herramientas eléctricas"),
            gettext("Proveedor de ejemplo"),
            "850.00",
            "1299.00",
            8,
            3,
        ],
        [
            "PIN-0001",
            "07501234560002",
            gettext("Pintura vinílica blanca 19 L"),
            gettext("Pinturas"),
            gettext("Proveedor de ejemplo"),
            "620.00",
            "899.00",
            12,
            4,
        ],
        [
            "TOR-0001",
            "00012345678905",
            gettext("Tornillo galvanizado 1/4 x 2"),
            gettext("Tornillería"),
            gettext("Proveedor de ejemplo"),
            "1.15",
            "2.50",
            500,
            100,
        ],
    ]
    examples = pd.DataFrame(example_rows, columns=columns)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = gettext("PRODUCTOS")
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=3)
        example_sheet_name = gettext("EJEMPLO FERRETERÍA")
        examples.to_excel(
            writer,
            index=False,
            sheet_name=example_sheet_name,
        )
        workbook = writer.book
        ws = writer.sheets[sheet_name]

        ws["A1"] = gettext("PATIA - Plantilla oficial de productos")
        ws["A2"] = gettext(
            "Llena esta tabla con tus productos. No cambies los nombres de las columnas."
        )
        ws.merge_cells("A1:I1")
        ws.merge_cells("A2:I2")

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo

        title_fill = PatternFill("solid", fgColor="0B1020")
        header_fill = PatternFill("solid", fgColor="00D4FF")
        dark_font = Font(color="0B1020", bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="D9E2F3")

        ws["A1"].fill = title_fill
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
        ws["A1"].alignment = center
        ws["A2"].font = Font(color="666666", italic=True)
        ws["A2"].alignment = center

        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = dark_font
            cell.alignment = center
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in range(5, 105):
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        table = Table(displayName="TablaProductosPATIA", ref="A4:I104")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        widths = {"A": 18, "B": 22, "C": 32, "D": 20, "E": 24, "F": 14, "G": 18, "H": 18, "I": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A5"

        example_ws = writer.sheets[example_sheet_name]
        for cell in example_ws[1]:
            cell.fill = header_fill
            cell.font = dark_font
            cell.alignment = center
        for col, width in widths.items():
            example_ws.column_dimensions[col].width = width
        example_ws.freeze_panes = "A2"

    output.seek(0)
    return send_file(output, as_attachment=True, download_name="plantilla_productos_PATIA.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@main.route("/import-products", methods=["POST"])
@require_permission("manage_inventory")
def import_products():
    import pandas as pd
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    access_block = _trial_access_response(user)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    membership = active_membership(user)
    owner = current_organization_owner(user)

    file = request.files.get("catalog_file")
    if not file:
        flash("Selecciona un archivo.", "danger")
        return redirect(url_for("main.products") + "#importar-catalogo")

    try:
        filename = (file.filename or "").lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
            first_data_row = 2
        elif filename.endswith(".xlsx"):
            workbook = pd.ExcelFile(file)
            sheet_name = next(
                (
                    name
                    for name in workbook.sheet_names
                    if name.upper() in {"PRODUCTOS", "PRODUCTS"}
                ),
                None,
            )
            if not sheet_name:
                raise ValueError("missing product sheet")
            df = pd.read_excel(workbook, sheet_name=sheet_name, header=3)
            first_data_row = 5
        else:
            flash("Usa un archivo CSV o Excel .xlsx.", "danger")
            return redirect(url_for("main.products") + "#importar-catalogo")

        column_aliases = {
            "SKU": "sku",
            "Codigo de barras": "barcode",
            "Código de barras": "barcode",
            "Barcode": "barcode",
            "Nombre del producto": "name",
            "Product name": "name",
            "Categoria": "category",
            "Categoría": "category",
            "Category": "category",
            "Proveedor": "supplier",
            "Supplier": "supplier",
            "Costo": "cost_price",
            "Cost": "cost_price",
            "Precio de venta": "sale_price",
            "Sale price": "sale_price",
            "Stock inicial": "stock",
            "Initial stock": "stock",
            "Stock minimo": "min_stock",
            "Stock mínimo": "min_stock",
            "Minimum stock": "min_stock",
        }
        df = df.rename(columns=column_aliases)
        essential_columns = {"sku", "name", "cost_price", "sale_price", "stock"}
        missing_columns = sorted(essential_columns - set(df.columns))
        if missing_columns:
            flash(
                gettext(
                    "El archivo no contiene las columnas obligatorias. Descarga la plantilla PATIA e inténtalo de nuevo."
                ),
                "danger",
            )
            return redirect(url_for("main.products") + "#importar-catalogo")

        summary = {
            "created": 0,
            "updated": 0,
            "omitted": 0,
            "errors": 0,
            "matched": 0,
        }
        existing_products = Product.query.filter_by(
            organization_id=organization_id
        ).all()
        products_by_sku = {
            product.sku: product for product in existing_products if product.sku
        }
        products_by_barcode = {
            product.barcode: product
            for product in existing_products
            if product.barcode
        }

        def text_value(value, default=""):
            if pd.isna(value):
                return default
            return str(value).strip()

        def number_value(value, default=0, integer=False):
            if pd.isna(value) or str(value).strip() == "":
                number = default
            elif not integer:
                return money_decimal(value)
            else:
                try:
                    number = Decimal(str(value))
                except (InvalidOperation, ValueError, TypeError) as error:
                    raise ValueError("invalid integer value") from error
            if integer:
                if not number.is_finite() or number < 0:
                    raise ValueError("negative value")
                if number != number.to_integral_value():
                    raise ValueError("fractional integer value")
                return int(number)
            return money_decimal(number)

        for row_index, row in df.iterrows():
            if all(pd.isna(value) or str(value).strip() == "" for value in row.values):
                summary["omitted"] += 1
                continue

            try:
                sku = text_value(row.get("sku"))
                name = text_value(row.get("name"))
                raw_barcode = row.get("barcode", "")
                try:
                    barcode = (
                        str(int(float(raw_barcode)))
                        if text_value(raw_barcode)
                        else ""
                    )
                except (TypeError, ValueError):
                    barcode = text_value(raw_barcode)

                stock = number_value(row.get("stock"), integer=True)
                cost_price = number_value(row.get("cost_price"))
                sale_price = number_value(row.get("sale_price"))
                min_stock = number_value(row.get("min_stock"), default=5, integer=True)

                existing = None
                matched_by_sku = False
                if sku:
                    existing = products_by_sku.get(sku)
                    matched_by_sku = existing is not None

                if not existing and barcode:
                    existing = products_by_barcode.get(barcode)

                if existing:
                    if matched_by_sku and barcode:
                        barcode_owner = products_by_barcode.get(barcode)
                        if barcode_owner and barcode_owner is not existing:
                            raise ValueError("barcode belongs to another product")
                    existing.is_active = True
                    # Política existente: SKU suma stock; código lo reemplaza.
                    stock_before = existing.stock
                    existing.stock = stock_before + stock if matched_by_sku else stock
                    existing.sale_price = sale_price
                    existing.cost_price = cost_price
                    existing.min_stock = min_stock
                    if matched_by_sku:
                        if existing.barcode:
                            products_by_barcode.pop(existing.barcode, None)
                        existing.barcode = barcode
                        if barcode:
                            products_by_barcode[barcode] = existing
                    record_inventory_movement(
                        existing,
                        membership,
                        "IMPORT",
                        stock_before,
                        existing.stock,
                        reason=(
                            gettext("Importación por coincidencia de SKU")
                            if matched_by_sku
                            else gettext("Importación por código de barras")
                        ),
                    )
                    summary["updated"] += 1
                    summary["matched"] += 1
                    continue

                if not sku or not name:
                    raise ValueError("missing product identity")

                imported_product = Product(
                    organization_id=organization_id,
                    user_id=owner.id,
                    sku=sku,
                    barcode=barcode or None,
                    name=name,
                    category=text_value(row.get("category"), "General") or "General",
                    supplier=text_value(row.get("supplier")) or None,
                    cost_price=cost_price,
                    sale_price=sale_price,
                    stock=stock,
                    min_stock=min_stock,
                )
                db.session.add(imported_product)
                products_by_sku[sku] = imported_product
                if barcode:
                    products_by_barcode[barcode] = imported_product
                record_opening_balance(
                    imported_product,
                    membership,
                    reason=gettext("Alta mediante importación"),
                )
                summary["created"] += 1
            except (TypeError, ValueError, OverflowError):
                summary["errors"] += 1
                current_app.logger.warning(
                    "Fila inválida en importación de catálogo (fila %s)",
                    row_index + first_data_row,
                    exc_info=True,
                )

        db.session.commit()
        flash(gettext(
            "Importación terminada: %(created)s creados, %(updated)s actualizados, %(omitted)s omitidos y %(errors)s errores. %(matched)s filas coincidieron con productos existentes.",
            **summary,
        ), "success")

    except Exception:
        db.session.rollback()
        current_app.logger.exception("No se pudo importar el catálogo")
        flash(
            "No pudimos procesar el archivo. Verifica que use la plantilla PATIA e inténtalo de nuevo.",
            "danger",
        )

    return redirect(url_for("main.products") + "#importar-catalogo")


def _catalog_upload():
    upload = request.files.get("catalog_file")
    if not upload or not upload.filename:
        raise ValueError("missing_file")
    content = upload.read()
    if len(content) > 12 * 1024 * 1024:
        raise ValueError("file_too_large")
    return upload.filename, content


def _catalog_error_message(code):
    messages = {
        "missing_file": gettext("Selecciona un archivo CSV o Excel."),
        "empty_file": gettext("El archivo está vacío."),
        "file_too_large": gettext("El archivo supera el límite de 12 MB."),
        "unsupported_file": gettext("Usa un archivo CSV o Excel .xlsx."),
        "too_many_rows": gettext(
            "El archivo supera el límite de %(limit)s productos.",
            limit=MAX_IMPORT_ROWS,
        ),
    }
    return messages.get(code, gettext("No pudimos leer el catálogo. Revisa el archivo e inténtalo nuevamente."))


def _catalog_row_error_message(code):
    messages = {
        "invalid_number": gettext("Revisa el costo o precio."),
        "ambiguous_number": gettext(
            "El formato del importe es ambiguo para la moneda del negocio. Usa separadores consistentes."
        ),
        "invalid_integer": gettext(
            "Stock y stock mínimo deben ser números enteros no negativos."
        ),
        "missing_identity": gettext("Agrega el nombre del producto."),
        "duplicate_in_file": gettext(
            "El SKU o código de barras está repetido dentro del archivo."
        ),
        "conflicting_identity": gettext(
            "El SKU y el código pertenecen a productos distintos."
        ),
    }
    return messages.get(
        code,
        gettext("Revisa los datos de esta fila."),
    )


@main.post("/api/products/import/preview")
@require_permission("manage_inventory")
def import_products_preview():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": gettext("Inicia sesión para continuar.")}), 401
    access_block = _trial_access_response(user, json_response=True)
    if access_block:
        return access_block
    try:
        filename, content = _catalog_upload()
        raw_mapping = request.form.get("mapping")
        mapping = json.loads(raw_mapping) if raw_mapping else None
        existing = Product.query.filter_by(
            organization_id=current_organization_id(user)
        ).all()
        membership = active_membership(user)
        currency_code, locale_code = organization_money_context(
            membership.organization
        )
        imported = inspect_catalog(
            filename, content, mapping, existing,
            currency_code=currency_code,
            locale_code=locale_code,
        )
        preview_rows = []
        for row in imported.rows[:20]:
            preview_rows.append({
                **row,
                "cost_price": str(row["cost_price"]),
                "sale_price": str(row["sale_price"]),
            })
        return jsonify({
            "ok": True,
            "filename": filename,
            "digest": imported.digest,
            "headers": imported.headers,
            "mapping": imported.mapping,
            "mapping_confidence": imported.mapping_confidence,
            "required_fields": ["name", "sale_price", "stock"],
            "summary": imported.summary,
            "rows": preview_rows,
            "errors": [
                {
                    **error,
                    "message": _catalog_row_error_message(error["code"]),
                }
                for error in imported.errors[:100]
            ],
            "ready": imported.summary["valid"] > 0 and all(
                key in imported.mapping
                for key in ("name", "sale_price", "stock")
            ),
        })
    except (ValueError, json.JSONDecodeError) as exc:
        return jsonify({"ok": False, "error": _catalog_error_message(str(exc))}), 400


@main.post("/api/products/import/commit")
@require_permission("manage_inventory")
def import_products_commit():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": gettext("Inicia sesión para continuar.")}), 401
    access_block = _trial_access_response(user, json_response=True)
    if access_block:
        return access_block
    try:
        filename, content = _catalog_upload()
        mapping = json.loads(request.form.get("mapping") or "{}")
        expected_digest = request.form.get("digest") or ""
        organization_id = current_organization_id(user)
        membership = active_membership(user)
        currency_code, locale_code = organization_money_context(
            membership.organization
        )
        imported = inspect_catalog(
            filename,
            content,
            mapping,
            Product.query.filter_by(organization_id=organization_id).all(),
            currency_code=currency_code,
            locale_code=locale_code,
        )
        if not secrets.compare_digest(imported.digest, expected_digest):
            raise ValueError("file_changed")
        if not imported.rows or not all(
            field in imported.mapping
            for field in ("name", "sale_price", "stock")
        ):
            return jsonify({
                "ok": False,
                "error": gettext("No encontramos filas válidas para importar."),
                "errors": [
                    {
                        **error,
                        "message": _catalog_row_error_message(error["code"]),
                    }
                    for error in imported.errors[:500]
                ],
            }), 409
        summary = apply_catalog(
            imported,
            organization_id,
            current_organization_owner(user).id,
            membership,
        )
        db.session.commit()
        return jsonify({
            "ok": True,
            "summary": summary,
            "message": gettext(
                "Catálogo listo: %(created)s creados, %(updated)s actualizados "
                "y %(errors)s rechazados.",
                **summary,
            ),
            "errors": [
                {
                    **error,
                    "message": _catalog_row_error_message(error["code"]),
                }
                for error in imported.errors
            ],
        })
    except (ValueError, json.JSONDecodeError) as exc:
        db.session.rollback()
        message = (
            gettext("El archivo cambió después de la vista previa. Vuelve a validarlo.")
            if str(exc) == "file_changed"
            else _catalog_error_message(str(exc))
        )
        return jsonify({"ok": False, "error": message}), 400
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": gettext("Otro proceso registró un SKU o código de barras durante la importación. Vuelve a validar el archivo."),
        }), 409
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "Error al confirmar importación para organization_id=%s",
            current_organization_id(user),
        )
        return jsonify({
            "ok": False,
            "error": gettext("No pudimos completar la importación. Ningún cambio parcial fue guardado."),
        }), 500


@main.route("/products/new", methods=["POST"])
@require_permission("manage_inventory")
def add_product():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    access_block = _trial_access_response(user)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    membership = active_membership(user)
    owner = current_organization_owner(user)

    name = request.form.get("name", "").strip()
    sku = request.form.get("sku", "").strip()
    try:
        cost_price = money_decimal(
            request.form.get("cost_price") or 0, nonnegative=False
        )
        sale_price = money_decimal(
            request.form.get("sale_price") or 0, nonnegative=False
        )
        stock = int(request.form.get("stock") or 0)
        min_stock = int(request.form.get("min_stock") or 5)
    except (TypeError, ValueError):
        flash("Revisa precios y existencias e inténtalo nuevamente.", "danger")
        return redirect(url_for("main.products"))
    if not name or not sku:
        flash("Nombre y SKU son obligatorios.", "danger")
        return redirect(url_for("main.products"))
    existing_sku = Product.query.filter_by(
        organization_id=organization_id, sku=sku
    ).with_for_update().first()
    if existing_sku and existing_sku.is_active:
        flash("Ya existe un producto con ese SKU. Usa un SKU diferente.", "danger")
        return redirect(url_for("main.products") + "#agregar-producto")
    barcode = request.form.get("barcode", "").strip() or None
    existing_barcode = (
        Product.query.filter_by(
            organization_id=organization_id, barcode=barcode
        ).with_for_update().first()
        if barcode else None
    )
    if existing_barcode and existing_barcode.is_active:
        flash("Ya existe un producto con ese código de barras.", "danger")
        return redirect(url_for("main.products") + "#agregar-producto")
    if existing_sku and existing_barcode and existing_sku.id != existing_barcode.id:
        flash(
            "El SKU y el código de barras pertenecen a productos históricos diferentes.",
            "danger",
        )
        return redirect(url_for("main.products") + "#agregar-producto")
    if (
        cost_price < MONEY_ZERO
        or sale_price < MONEY_ZERO
        or stock < 0
        or min_stock < 0
    ):
        flash("Precios y existencias no pueden ser negativos.", "danger")
        return redirect(url_for("main.products"))

    p = existing_sku or existing_barcode
    stock_before = p.stock if p else 0
    if p:
        p.sku = sku
        p.barcode = barcode
        p.name = name
        p.category = request.form.get("category") or "General"
        p.supplier = request.form.get("supplier")
        p.cost_price = cost_price
        p.sale_price = sale_price
        p.stock = stock
        p.min_stock = min_stock
        p.is_active = True
    else:
        p = Product(
            organization_id=organization_id,
            user_id=owner.id,
            sku=sku,
            barcode=barcode,
            name=name,
            category=request.form.get("category") or "General",
            supplier=request.form.get("supplier"),
            cost_price=cost_price,
            sale_price=sale_price,
            stock=stock,
            min_stock=min_stock,
        )
        db.session.add(p)
    try:
        if p.id:
            record_inventory_movement(
                p,
                membership,
                "PHYSICAL_COUNT",
                stock_before,
                p.stock,
                reason=gettext("Restauración de producto archivado"),
            )
        else:
            record_opening_balance(
                p,
                membership,
                reason=gettext("Alta manual de producto"),
            )
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        current_app.logger.info(
            "Alta de producto rechazada por identificador duplicado para organization_id=%s",
            organization_id,
        )
        flash("No pudimos guardar el producto porque el SKU ya está en uso.", "danger")
        return redirect(url_for("main.products") + "#agregar-producto")
    flash("Producto creado correctamente.", "success")
    return redirect(url_for("main.products"))


@main.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
@require_permission("manage_inventory")
def edit_product(product_id):
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    access_block = _trial_access_response(user)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    membership = active_membership(user)
    product = Product.query.filter_by(
        id=product_id,
        organization_id=organization_id,
        is_active=True,
    ).with_for_update().first_or_404()
    if request.method == "GET":
        return render_template("edit_product.html", product=product, user=user)

    name = request.form.get("name", "").strip()
    sku = request.form.get("sku", "").strip()
    barcode = request.form.get("barcode", "").strip() or None
    try:
        cost_price = money_decimal(
            request.form.get("cost_price") or 0, nonnegative=False
        )
        sale_price = money_decimal(
            request.form.get("sale_price") or 0, nonnegative=False
        )
        stock = int(request.form.get("stock") or 0)
        min_stock = int(request.form.get("min_stock") or 0)
    except (TypeError, ValueError):
        flash("Revisa precios y existencias e inténtalo nuevamente.", "danger")
        return render_template("edit_product.html", product=product, user=user), 400

    if not name or not sku:
        flash("Nombre y SKU son obligatorios.", "danger")
        return render_template("edit_product.html", product=product, user=user), 400
    if (
        cost_price < MONEY_ZERO
        or sale_price < MONEY_ZERO
        or stock < 0
        or min_stock < 0
    ):
        flash("Precios y existencias no pueden ser negativos.", "danger")
        return render_template("edit_product.html", product=product, user=user), 400

    duplicate_sku = Product.query.filter(
        Product.organization_id == organization_id,
        Product.sku == sku,
        Product.id != product.id,
    ).first()
    if duplicate_sku:
        flash("Ya existe otro producto con ese SKU.", "danger")
        return render_template("edit_product.html", product=product, user=user), 409

    if barcode:
        duplicate_barcode = Product.query.filter(
            Product.organization_id == organization_id,
            Product.barcode == barcode,
            Product.id != product.id,
        ).first()
        if duplicate_barcode:
            flash("Ya existe otro producto con ese código de barras.", "danger")
            return render_template("edit_product.html", product=product, user=user), 409

    stock_before = product.stock
    product.name = name
    product.sku = sku
    product.barcode = barcode
    product.category = request.form.get("category", "").strip() or "General"
    product.supplier = request.form.get("supplier", "").strip() or None
    product.cost_price = cost_price
    product.sale_price = sale_price
    product.stock = stock
    product.min_stock = min_stock
    try:
        if stock_before != product.stock:
            record_inventory_movement(
                product,
                membership,
                "PHYSICAL_COUNT",
                stock_before,
                product.stock,
                reason=gettext("Stock actualizado al editar el producto"),
            )
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        current_app.logger.info(
            "Edición de producto rechazada por identificador duplicado para organization_id=%s",
            organization_id,
        )
        flash("No pudimos guardar el producto porque el SKU ya está en uso.", "danger")
        return render_template("edit_product.html", product=product, user=user), 409

    flash("Producto actualizado correctamente. Las ventas anteriores conservaron sus importes originales.", "success")
    return redirect(url_for("main.products") + "#catalogo")


@main.route("/products/<int:product_id>/restock", methods=["POST"])
@require_permission("make_inventory_adjustments")
def restock_product(product_id):
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    access_block = _trial_access_response(user)
    if access_block:
        return access_block

    organization_id = current_organization_id(user)
    membership = active_membership(user)
    owner = current_organization_owner(user)

    try:
        quantity = int(request.form.get("quantity", ""))
    except (TypeError, ValueError):
        flash(gettext("Ingresa una cantidad recibida válida."), "danger")
        return redirect(url_for("main.dashboard"))
    if quantity <= 0:
        flash(gettext("La cantidad recibida debe ser mayor que cero."), "danger")
        return redirect(url_for("main.dashboard"))

    product = (
        Product.query.filter_by(
            id=product_id,
            organization_id=organization_id,
            is_active=True,
        )
        .with_for_update()
        .first_or_404()
    )
    try:
        if product.stock > 2_147_483_647 - quantity:
            flash(gettext("La cantidad recibida supera el límite permitido."), "danger")
            db.session.rollback()
            return redirect(url_for("main.dashboard"))

        stock_before = product.stock
        product.stock = stock_before + quantity
        restock_event = InventoryRestockEvent(
                organization_id=organization_id,
                user_id=owner.id,
                product_id=product.id,
                quantity=quantity,
                stock_before=stock_before,
                stock_after=product.stock,
        )
        db.session.add(restock_event)
        db.session.flush()
        record_inventory_movement(
            product,
            membership,
            "RESTOCK",
            stock_before,
            product.stock,
            reason=gettext("Reabastecimiento de inventario"),
            restock_event=restock_event,
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "No se pudo registrar el reabastecimiento para organization_id=%s product_id=%s",
            organization_id,
            product_id,
        )
        flash(
            gettext("No pudimos actualizar el inventario. Inténtalo nuevamente."),
            "danger",
        )
        return redirect(url_for("main.dashboard"))

    flash(
        gettext(
            "Se agregaron %(quantity)s unidades a %(product)s.",
            quantity=quantity,
            product=product.name,
        ),
        "success",
    )
    return redirect(url_for("main.dashboard"))


@main.route("/sell", methods=["GET", "POST"])
@require_permission("use_pos")
def sell():
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    organization_id = current_organization_id(user)
    membership = active_membership(user)
    cash_session = open_cash_session(organization_id)
    owner = current_organization_owner(user)
    if trial_expired(user):
        return render_template("trial_expired.html")

    if request.method == "POST":
        try:
            product_id = int(request.form.get("product_id", ""))
            qty = int(request.form.get("quantity") or 1)
        except (TypeError, ValueError):
            flash("Selecciona un producto y una cantidad válida.", "danger")
            return redirect(url_for("main.sell"))
        product = (
            Product.query.filter_by(
                id=product_id,
                organization_id=organization_id,
                is_active=True,
            )
            .with_for_update()
            .first_or_404()
        )
        if qty <= 0:
            flash("La cantidad debe ser mayor a cero.", "danger")
        elif product.stock < qty:
            flash("No hay suficiente inventario.", "danger")
        else:
            payment_method = request.form.get("payment_method", "cash")
            if payment_method not in PAYMENT_METHOD_LABELS:
                flash("Selecciona un método de pago válido.", "danger")
                return redirect(url_for("main.sell"))
            if payment_method == "cash" and not cash_session:
                flash(
                    "Abre la caja antes de registrar una venta en efectivo.",
                    "danger",
                )
                return redirect(url_for("cash.index"))
            try:
                customer = _selected_customer(
                    organization_id,
                    request.form.get("customer_id"),
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return redirect(url_for("main.sell"))
            ticket = _create_sales_ticket(user, payment_method)
            ticket.customer_id = customer.id if customer else None
            ticket.cash_register_session_id = (
                cash_session.id if cash_session else None
            )
            stock_before = product.stock
            product.stock -= qty
            sale = Sale(
                organization_id=organization_id,
                user_id=owner.id,
                product_id=product.id,
                quantity=qty,
                unit_price=product.sale_price,
                unit_cost=product.cost_price if product.cost_price > 0 else None,
                cost_is_estimated=False,
                total=qty * product.sale_price,
                ticket_id=ticket.public_id,
                sales_ticket_id=ticket.id,
                payment_method=payment_method,
                currency_code=ticket.currency_code,
                locale_code=ticket.locale_code,
            )
            db.session.add(sale)
            db.session.flush()
            if payment_method == "credit":
                if not has_permission(membership, "use_customer_credit"):
                    db.session.rollback()
                    abort(403)
                if not customer:
                    db.session.rollback()
                    flash("Selecciona un cliente para vender a crédito.", "danger")
                    return redirect(url_for("main.sell"))
                try:
                    record_credit_charge(
                        customer,
                        membership,
                        sale.total,
                        ticket,
                        allow_override=request.form.get("credit_override") == "1",
                        override_pin=request.form.get("override_pin"),
                    )
                except CreditError as exc:
                    db.session.rollback()
                    flash(str(exc), "danger")
                    return redirect(url_for("main.sell"))
            record_inventory_movement(
                product,
                membership,
                "SALE",
                stock_before,
                product.stock,
                reason=gettext("Venta registrada"),
                sale=sale,
                sales_ticket=ticket,
            )
            if payment_method == "cash":
                record_cash_movement(
                    cash_session,
                    membership,
                    "SALE_CASH",
                    sale.total,
                    note=ticket.folio,
                    sales_ticket=ticket,
                )
            db.session.commit()
            flash(
                gettext(
                    "Venta registrada: %(product)s x%(quantity)s.",
                    product=product.name,
                    quantity=qty,
                ),
                "success",
            )
        return redirect(url_for("main.sell"))

    sales = (
        Sale.query.options(
            selectinload(Sale.product),
            selectinload(Sale.sales_ticket),
        )
        .filter_by(organization_id=organization_id)
        .order_by(Sale.created_at.desc(), Sale.id.desc())
        .all()
    )
    sale_groups = _group_sales_by_ticket(
        sales,
        limit=12,
        timezone_name=active_membership(user).organization.timezone,
    )
    products = Product.query.filter_by(
        organization_id=organization_id,
        is_active=True,
    ).order_by(Product.name).all()
    return render_template(
        "sell.html",
        products=products,
        sales=sales,
        sale_groups=sale_groups,
        user=user,
        payment_method_labels=_translated_payment_method_labels(),
        cash_session=cash_session,
        cashier_mode=bool(session.get("cashier_mode")),
    )


@main.post("/sell/cashier-mode")
@require_permission("use_pos")
def sell_cashier_mode():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    session["cashier_mode"] = request.form.get("enabled") == "1"
    session.modified = True
    return redirect(url_for("main.sell"))


@main.route("/sell-cart", methods=["POST"])
@limiter.limit(
    "5 per minute",
    methods=["POST"],
    key_func=_credit_override_rate_limit_key,
    exempt_when=_credit_override_rate_limit_exempt,
    on_breach=_credit_override_rate_limit_response,
)
@require_permission("use_pos")
def sell_cart():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": gettext("No autenticado")}), 401
    access_block = _trial_access_response(user, json_response=True)
    if access_block:
        return access_block
    organization_id = current_organization_id(user)
    membership = active_membership(user)
    owner = current_organization_owner(user)

    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"ok": False, "error": gettext("Solicitud inválida")}), 400

    items = data.get("items")
    if not isinstance(items, list) or not items:
        return jsonify({"ok": False, "error": gettext("El carrito está vacío")}), 400

    payment_method = data.get("payment_method", "cash")
    if payment_method not in PAYMENT_METHOD_LABELS:
        return jsonify({"ok": False, "error": gettext("Selecciona un método de pago válido")}), 400
    cash_session = open_cash_session(organization_id, lock=True)
    if payment_method == "cash" and not cash_session:
        return jsonify({
            "ok": False,
            "error": gettext(
                "Abre la caja antes de registrar una venta en efectivo."
            ),
            "cash_register_url": url_for("cash.index"),
        }), 409
    amount_received = None
    if payment_method == "cash" and data.get("amount_received") in (None, ""):
        return jsonify({
            "ok": False,
            "error": gettext("Ingresa el efectivo recibido para continuar."),
        }), 400
    if payment_method == "cash":
        try:
            amount_received = money_decimal(data.get("amount_received"))
        except (TypeError, ValueError):
            return jsonify({
                "ok": False,
                "error": gettext("Ingresa un monto recibido válido."),
            }), 400
        if amount_received < 0:
            return jsonify({
                "ok": False,
                "error": gettext("El monto recibido no puede ser negativo."),
            }), 400
    try:
        customer = _selected_customer(
            organization_id,
            data.get("customer_id"),
        )
    except ValueError as exc:
        return jsonify({"ok": False, "error": gettext(str(exc))}), 400

    request_id = data.get("request_id")
    if request_id:
        try:
            request_id = str(uuid.UUID(str(request_id)))
        except (TypeError, ValueError, AttributeError):
            return jsonify({"ok": False, "error": gettext("Solicitud inválida")}), 400

    try:
        requested_items = {}
        for item in items:
            if not isinstance(item, dict):
                return jsonify({"ok": False, "error": gettext("Artículo inválido")}), 400

            try:
                product_id = int(item["product_id"])
                quantity = int(item["quantity"])
            except (KeyError, TypeError, ValueError):
                return jsonify({"ok": False, "error": gettext("Artículo inválido")}), 400

            if quantity <= 0:
                return jsonify({"ok": False, "error": gettext("La cantidad debe ser mayor a cero")}), 400

            requested_items[product_id] = requested_items.get(product_id, 0) + quantity

        if request_id:
            previous_sales = Sale.query.filter_by(
                organization_id=organization_id,
                ticket_id=request_id,
            ).order_by(Sale.id).all()
            if previous_sales:
                folio = _short_sale_folio(previous_sales)
                return jsonify({
                    "ok": True,
                    "duplicate": True,
                    "ticket_id": request_id,
                    "folio": folio,
                    "ticket_url": url_for("main.ticket", ticket_ref=request_id),
                    "total": money_json(money_sum(sale.total for sale in previous_sales)),
                    "payment_method": _payment_method_label(previous_sales[0].payment_method),
                    "single_sale_id": (
                        previous_sales[0].id if len(previous_sales) == 1 else None
                    ),
                })

        products = {
            product.id: product
            for product in Product.query.filter(
                Product.organization_id == organization_id,
                Product.is_active.is_(True),
                Product.id.in_(requested_items.keys()),
            )
            .order_by(Product.id)
            .with_for_update()
            .all()
        }

        if len(products) != len(requested_items):
            return jsonify({"ok": False, "error": gettext("Producto no encontrado")}), 404

        expected_total = money_sum(
            products[product_id].sale_price * quantity
            for product_id, quantity in requested_items.items()
        )
        if payment_method == "cash":
            if amount_received < expected_total:
                return jsonify({
                    "ok": False,
                    "error": gettext(
                        "Faltan %(amount)s para completar el pago.",
                        amount=money_json(expected_total - amount_received),
                    ),
                }), 400

        # Repetir la verificación tras bloquear inventario evita que dos workers
        # procesen simultáneamente el mismo request_id.
        if request_id:
            previous_sales = Sale.query.filter_by(
                organization_id=organization_id,
                ticket_id=request_id,
            ).order_by(Sale.id).all()
            if previous_sales:
                return jsonify({
                    "ok": True,
                    "duplicate": True,
                    "ticket_id": request_id,
                    "folio": _short_sale_folio(previous_sales),
                    "ticket_url": url_for("main.ticket", ticket_ref=request_id),
                    "total": money_json(money_sum(sale.total for sale in previous_sales)),
                    "payment_method": _payment_method_label(previous_sales[0].payment_method),
                    "single_sale_id": (
                        previous_sales[0].id if len(previous_sales) == 1 else None
                    ),
                })

        for product_id, quantity in requested_items.items():
            product = products[product_id]
            if product.stock < quantity:
                return jsonify({
                    "ok": False,
                    "error": gettext("Stock insuficiente: %(product)s", product=product.name),
                }), 409

        ticket = _create_sales_ticket(
            user,
            payment_method,
            ticket_id=request_id,
        )
        ticket.cash_register_session_id = (
            cash_session.id if cash_session else None
        )
        ticket.customer_id = customer.id if customer else None
        if payment_method == "cash":
            ticket.amount_received = amount_received
            ticket.change_amount = amount_received - expected_total
        ticket_id = ticket.public_id
        sales = []
        for product_id, quantity in requested_items.items():
            product = products[product_id]
            stock_before = product.stock
            product.stock -= quantity
            sale = Sale(
                organization_id=organization_id,
                user_id=owner.id,
                product_id=product.id,
                quantity=quantity,
                unit_price=product.sale_price,
                unit_cost=product.cost_price if product.cost_price > 0 else None,
                cost_is_estimated=False,
                total=quantity * product.sale_price,
                payment_method=payment_method,
                currency_code=ticket.currency_code,
                locale_code=ticket.locale_code,
                sales_ticket_id=ticket.id,
            )
            sale.ticket_id = ticket_id
            db.session.add(sale)
            db.session.flush()
            record_inventory_movement(
                product,
                membership,
                "SALE",
                stock_before,
                product.stock,
                reason=gettext("Venta registrada"),
                sale=sale,
                sales_ticket=ticket,
            )
            sales.append(sale)
        db.session.flush()
        if payment_method == "credit":
            if not has_permission(membership, "use_customer_credit"):
                db.session.rollback()
                return jsonify(
                    {"ok": False, "error": gettext("Acceso no permitido.")}
                ), 403
            if not customer:
                raise CreditError(
                    gettext("Selecciona un cliente para vender a crédito.")
                )
            record_credit_charge(
                customer,
                membership,
                money_sum(sale.total for sale in sales),
                ticket,
                allow_override=bool(data.get("credit_override")),
                override_pin=data.get("override_pin"),
            )
        if payment_method == "cash":
            record_cash_movement(
                cash_session,
                membership,
                "SALE_CASH",
                money_sum(sale.total for sale in sales),
                note=ticket.folio,
                sales_ticket=ticket,
            )
        db.session.commit()
        folio = _short_sale_folio(sales)
        return jsonify({
            "ok": True,
            "ticket_id": ticket_id,
            "folio": folio,
            "ticket_url": url_for("main.ticket", ticket_ref=ticket_id),
            "total": money_json(money_sum(sale.total for sale in sales)),
            "single_sale_id": sales[0].id if len(sales) == 1 else None,
            "payment_method": _payment_method_label(payment_method),
        })
    except CreditNotEnabled as exc:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(exc),
            "error_code": "credit_not_enabled",
        }), 409
    except CreditLimitExceeded as exc:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(exc),
            "error_code": "credit_limit_exceeded",
            "projected_balance": money_json(exc.balance),
            "credit_limit": money_json(exc.limit),
        }), 409
    except CreditError as exc:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(exc)}), 409
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Error al procesar el carrito")
        return jsonify({"ok": False, "error": gettext("No se pudo procesar la venta")}), 500


@main.route("/reports")
@require_permission("view_reports")
def reports():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    owner = current_organization_owner(user)
    membership = active_membership(user)
    from .plans import has_entitlement

    advanced_reports = has_entitlement(owner, "advanced_reports")
    requested_period = request.args.get("period", "7d")
    period_args = request.args
    if (
        not advanced_reports
        and requested_period
        in {"30d", "previous_month", "custom"}
    ):
        period_args = {"period": "7d"}
        flash(
            gettext(
                "Los periodos y análisis avanzados están incluidos en PATIA Pro."
            ),
            "info",
        )
    timezone_name = safe_timezone_name(membership.organization.timezone)
    report_period = _parse_report_period(
        period_args,
        timezone_name=timezone_name,
    )
    analytics = _report_analytics(
        membership.organization_id,
        report_period,
        timezone_name=timezone_name,
        currency_code=membership.organization.currency_code,
        include_mixed_currency=True,
    )
    executive_data = None
    if advanced_reports:
        from .pro.services import build_executive_dashboard

        executive_data = build_executive_dashboard(
            membership.organization,
            {
                "period": report_period["period"],
                "start": report_period["custom_start"],
                "end": report_period["custom_end"],
            },
        )
    client_daily_report = analytics["daily_report"]
    if not advanced_reports:
        client_daily_report = [
            {"date": point["date"], "sales": point["sales"]}
            for point in client_daily_report
        ]
    return render_template(
        "reports.html",
        user=user,
        can_use_advanced_reports=advanced_reports,
        can_edit_goal=has_permission(membership, "manage_subscription"),
        client_daily_report=client_daily_report,
        executive_data=executive_data,
        **analytics,
    )


@main.route("/suppliers", methods=["GET", "POST"])
@require_permission("manage_inventory")
def suppliers():
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    organization_id = current_organization_id(user)
    owner = current_organization_owner(user)

    if request.method == "POST":
        if not has_permission(active_membership(user), "manage_inventory"):
            abort(403)
        access_block = _trial_access_response(user)
        if access_block:
            return access_block
        supplier_name = request.form.get("name", "").strip()
        if not supplier_name:
            flash("Escribe el nombre del proveedor.", "danger")
            return redirect(url_for("main.suppliers"))
        existing_supplier = Supplier.query.filter_by(
            organization_id=organization_id, name=supplier_name
        ).first()
        if existing_supplier:
            flash("Ese proveedor ya existe.", "danger")
            return redirect(url_for("main.suppliers"))
        s = Supplier(
            organization_id=organization_id,
            user_id=owner.id,
            name=supplier_name,
            contact=request.form.get("contact"),
            phone=request.form.get("phone"),
            notes=request.form.get("notes"),
        )
        db.session.add(s)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Ese proveedor ya existe.", "danger")
            return redirect(url_for("main.suppliers"))
        flash("Proveedor guardado.", "success")
        return redirect(url_for("main.suppliers"))

    suppliers = Supplier.query.filter_by(
        organization_id=organization_id
    ).order_by(Supplier.name).all()
    return render_template("suppliers.html", suppliers=suppliers, user=user)


@main.route("/subscribe")
@require_permission("manage_subscription")
def subscribe():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    if not user.email_verified:
        session["post_verify_destination"] = "subscribe"
        flash("Verifica tu correo antes de activar PATIA Pro.", "info")
        return redirect(url_for("main.verify_email"))
    if request.args.get("checkout") == "cancelled":
        flash("El pago fue cancelado. No se realizó ningún cargo.", "info")
    from .plans import commercial_plans

    return render_template(
        "subscribe.html",
        user=user,
        commercial_plans=commercial_plans(current_app.config),
    )


def _safe_stripe_error(error):
    """Return diagnostic fields without leaking Stripe identifiers or URLs."""
    message = str(error or "")[:1000]
    message = re.sub(r"https?://\S+", "[redacted-url]", message)
    message = re.sub(
        r"\b(?:sk|rk|pk|whsec|cs|sess|pi|sub|cus|price|prod|req)_[A-Za-z0-9_]+\b",
        "[redacted-stripe-id]",
        message,
    )


    code = getattr(error, "code", None) or getattr(error, "http_status", None)
    return type(error).__name__, code or "none", message or "none"


@main.route("/terminos")
def terms():
    return render_template("legal.html", document="terms")


@main.route("/privacidad")
def privacy():
    return render_template("legal.html", document="privacy")


@main.route("/create-checkout-session", methods=["POST"])
@require_permission("manage_subscription")
def create_checkout_session():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    user = current_organization_owner(user)
    if not user.email_verified:
        session["post_verify_destination"] = "subscribe"
        flash("Verifica tu correo antes de activar PATIA Pro.", "info")
        return redirect(url_for("main.verify_email"))
    if current_app.config["STRIPE_DISABLED"]:
        flash("La facturación no está disponible en este entorno.", "danger")
        return redirect(url_for("main.subscribe"))

    from .plans import PAID_PLAN_CODES, STARTER, price_id_for

    requested_plan = str(
        request.form.get("plan_code") or STARTER
    ).strip().upper()
    if requested_plan not in PAID_PLAN_CODES:
        flash("Selecciona un plan válido.", "danger")
        return redirect(url_for("main.subscribe"))
    price_id = price_id_for(current_app.config, requested_plan)
    if not price_id:
        flash(
            "Este plan estará disponible para contratación muy pronto.",
            "info",
        )
        return redirect(url_for("main.subscribe"))

    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    existing_subscription = None

    if (
        user.stripe_subscription_id
        and user.subscription_status in MANAGED_SUBSCRIPTION_STATUSES
    ):
        return _redirect_to_billing_portal(user)

    if user.stripe_customer_id:
        try:
            subscriptions = stripe.Subscription.list(
                customer=user.stripe_customer_id,
                status="all",
                limit=10,
            )
            for candidate in subscriptions.get("data", []):
                if (
                    candidate.get("status") in MANAGED_SUBSCRIPTION_STATUSES
                    and _subscription_has_configured_price(candidate)
                ):
                    existing_subscription = candidate
                    break
        except Exception as error:
            error_type, error_code, error_message = _safe_stripe_error(error)
            current_app.logger.exception(
                "No se pudo comprobar suscripciones existentes: type=%s code=%s message=%s",
                error_type,
                error_code,
                error_message,
            )
            flash("No pudimos validar tu suscripción. Intenta nuevamente.", "danger")
            return redirect(url_for("main.subscribe"))

    if existing_subscription:
        user.stripe_subscription_id = existing_subscription.get("id")
        _sync_subscription_state(user, existing_subscription, datetime.utcnow())
        db.session.commit()
        return _redirect_to_billing_portal(user)

    checkout_params = {
        "payment_method_types": ["card"],
        "mode": "subscription",
        "line_items": [{"price": price_id, "quantity": 1}],
        "success_url": _public_url("/stripe-success") + "?session_id={CHECKOUT_SESSION_ID}",
        "cancel_url": _public_url("/subscribe?checkout=cancelled"),
        "client_reference_id": str(user.id),
        "metadata": {
            "user_id": str(user.id),
            "organization_id": str(current_organization_id(user)),
            "plan_code": requested_plan,
        },
        "subscription_data": {
            "metadata": {
                "user_id": str(user.id),
                "organization_id": str(current_organization_id(user)),
                "plan_code": requested_plan,
            }
        },
    }
    if user.stripe_customer_id:
        checkout_params["customer"] = user.stripe_customer_id
    else:
        checkout_params["customer_email"] = user.email

    idempotency_window = int(datetime.utcnow().timestamp() // 1800)
    try:
        checkout_session = stripe.checkout.Session.create(
            **checkout_params,
            idempotency_key=(
                f"patia-checkout-{user.id}-{requested_plan.lower()}-"
                f"{idempotency_window}"
            ),
        )
        return redirect(checkout_session.url, code=303)
    except Exception as error:
        error_type, error_code, error_message = _safe_stripe_error(error)
        current_app.logger.exception(
            "No se pudo crear la sesión de Checkout: type=%s code=%s message=%s",
            error_type,
            error_code,
            error_message,
        )
        flash("No pudimos iniciar el pago. Intenta nuevamente.", "danger")
        return redirect(url_for("main.subscribe"))


def _process_stripe_event(event):
    event_type = event["type"]
    data = event["data"]["object"]
    stripe_created_at = _as_utc_datetime(event["created"])

    if event_type == "checkout.session.completed":
        if data.get("mode") != "subscription":
            raise StripeEventIgnored("Checkout no es de suscripción.")
        user_id = str(
            data.get("client_reference_id")
            or (data.get("metadata") or {}).get("user_id")
            or ""
        )
        if not user_id.isdigit():
            raise StripeEventIgnored("Checkout sin usuario válido.")
        user = db.session.get(User, int(user_id))
        if not user:
            raise StripeEventIgnored("Usuario de Checkout inexistente.")
        customer_id = data.get("customer")
        subscription_id = data.get("subscription")
        if not customer_id or not subscription_id:
            raise StripeEventIgnored("Checkout sin cliente o suscripción.")
        if user.stripe_customer_id and user.stripe_customer_id != customer_id:
            raise StripeEventIgnored("Checkout pertenece a otro cliente.")
        _ensure_stripe_ids_available(user, customer_id, subscription_id)
        subscription = stripe.Subscription.retrieve(subscription_id)
        _validate_subscription(subscription, user=user, customer_id=customer_id)
        metadata_user_id = str((subscription.get("metadata") or {}).get("user_id") or "")
        if metadata_user_id and metadata_user_id != str(user.id):
            raise StripeEventIgnored("Metadatos de suscripción no coinciden.")
        metadata_org_id = str(
            (subscription.get("metadata") or {}).get("organization_id") or ""
        )
        organization = Organization.query.filter_by(owner_user_id=user.id).first()
        if (
            metadata_org_id
            and organization
            and metadata_org_id != str(organization.id)
        ):
            raise StripeEventIgnored(
                "Metadatos de organización no coinciden."
            )
        requested_plan = str(
            (data.get("metadata") or {}).get("plan_code") or ""
        ).upper()
        subscription_plan = _subscription_plan_code(subscription)
        if requested_plan and requested_plan != subscription_plan:
            raise StripeEventIgnored(
                "El plan de Checkout no coincide con la suscripción."
            )
        user.stripe_customer_id = customer_id
        user.stripe_subscription_id = subscription_id
        return

    if event_type in {"invoice.paid", "invoice.payment_failed"}:
        subscription_id = _invoice_subscription_id(data)
        if not subscription_id:
            raise StripeEventIgnored("Factura sin suscripción.")
        subscription = stripe.Subscription.retrieve(subscription_id)
        _validate_subscription(subscription, customer_id=data.get("customer"))
        user = _find_subscription_user(subscription)
        if not user:
            raise StripeEventIgnored("Suscripción sin usuario local.")
        _validate_subscription(subscription, user=user, customer_id=data.get("customer"))
        if not _event_is_newer(user, stripe_created_at, "invoice"):
            return
        subscription_event_is_newer = bool(
            user.stripe_subscription_updated_at
            and stripe_created_at < user.stripe_subscription_updated_at
        )
        user.stripe_customer_id = subscription.get("customer")
        user.stripe_subscription_id = subscription_id
        user.current_period_end = _subscription_period_end(subscription)
        user.cancel_at_period_end = bool(subscription.get("cancel_at_period_end", False))
        user.stripe_invoice_updated_at = stripe_created_at
        if event_type == "invoice.paid":
            status = (subscription.get("status") or "").lower()
            if status not in {"active", "trialing"}:
                raise StripeEventIgnored("Factura pagada con suscripción no activa.")
            if not subscription_event_is_newer:
                user.subscription_status = status
                user.next_payment_attempt = None
            resolved_plan = _subscription_plan_code(
                subscription,
                preserve_legacy=not (
                    user.subscription_plan_code
                    or (subscription.get("metadata") or {}).get("plan_code")
                ),
            )
            user.subscription_plan_code = resolved_plan
            if user.pending_plan_code == resolved_plan:
                user.pending_plan_code = None
                user.pending_plan_effective_at = None
        else:
            if not subscription_event_is_newer:
                user.subscription_status = "past_due"
                user.next_payment_attempt = _as_utc_datetime(
                    data.get("next_payment_attempt")
                )
        sync_user_plan(user)
        return

    if event_type in {"customer.subscription.updated", "customer.subscription.deleted"}:
        _validate_subscription(data)
        user = _find_subscription_user(data)
        if not user:
            raise StripeEventIgnored("Suscripción sin usuario local.")
        _validate_subscription(data, user=user)
        _sync_subscription_state(
            user,
            data,
            stripe_created_at,
            deleted=event_type == "customer.subscription.deleted",
        )


def _record_failed_webhook(event, error):
    event_id = str(event.get("id") or "")
    if not event_id:
        return
    record = StripeWebhookEvent.query.filter_by(stripe_event_id=event_id).first()
    if not record:
        data = (event.get("data") or {}).get("object") or {}
        record = StripeWebhookEvent(
            stripe_event_id=event_id,
            event_type=str(event.get("type") or "unknown"),
            object_id=data.get("id"),
            stripe_created_at=_as_utc_datetime(event.get("created")) or datetime.utcnow(),
        )
        db.session.add(record)
    record.status = "failed"
    record.completed_at = None
    record.failed_at = datetime.utcnow()
    record.error_message = str(error)[:1000]
    db.session.commit()


@main.route("/stripe-webhook", methods=["POST"])
@csrf.exempt
def stripe_webhook():
    payload = request.data
    sig_header = request.headers.get("Stripe-Signature")
    endpoint_secret = current_app.config["STRIPE_WEBHOOK_SECRET"]

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
    except ValueError:
        return "", 400
    except stripe.error.SignatureVerificationError:
        return "", 400

    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    event_id = str(event.get("id") or "")
    if not event_id:
        return "", 400

    existing = StripeWebhookEvent.query.filter_by(stripe_event_id=event_id).first()
    if existing and existing.status in {"processed", "ignored"}:
        return "", 200

    data = event["data"]["object"]
    try:
        if not existing:
            existing = StripeWebhookEvent(
                stripe_event_id=event_id,
                event_type=event["type"],
                object_id=data.get("id"),
                stripe_created_at=_as_utc_datetime(event["created"]),
                status="pending",
            )
            db.session.add(existing)
            db.session.flush()
        else:
            existing.status = "pending"
            existing.error_message = None
            existing.completed_at = None
            existing.failed_at = None

        _process_stripe_event(event)
        existing.status = "processed"
        existing.completed_at = datetime.utcnow()
        existing.failed_at = None
        db.session.commit()
        return "", 200
    except StripeEventIgnored as error:
        existing.status = "ignored"
        existing.error_message = str(error)[:1000]
        existing.completed_at = datetime.utcnow()
        existing.failed_at = None
        db.session.commit()
        current_app.logger.warning("Evento Stripe ignorado: %s", error)
        return "", 200
    except IntegrityError:
        db.session.rollback()
        duplicate = StripeWebhookEvent.query.filter_by(stripe_event_id=event_id).first()
        if duplicate and duplicate.status in {"processed", "ignored"}:
            return "", 200
        current_app.logger.exception("Conflicto procesando webhook Stripe")
        return "", 500
    except Exception as error:
        db.session.rollback()
        current_app.logger.exception("Error procesando webhook Stripe %s", event_id)
        try:
            _record_failed_webhook(event, error)
        except Exception:
            db.session.rollback()
            current_app.logger.exception("No se pudo registrar el fallo del webhook")
        return "", 500


@main.route("/stripe-success")
@require_permission("manage_subscription")
def stripe_success():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    user = current_organization_owner(user)

    checkout_session_id = request.args.get("session_id", "").strip()
    if not checkout_session_id:
        flash("No se pudo validar la sesión de pago.", "danger")
        return redirect(url_for("main.subscription"))

    try:
        stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
        checkout_session = stripe.checkout.Session.retrieve(checkout_session_id)
    except Exception:
        current_app.logger.exception("No se pudo consultar la sesión de Checkout")
        flash("No se pudo validar la sesión de pago.", "danger")
        return redirect(url_for("main.subscription"))

    checkout_user_id = str(
        checkout_session.get("client_reference_id")
        or checkout_session.get("metadata", {}).get("user_id")
        or ""
    )
    if checkout_user_id != str(user.id):
        flash("La sesión de pago no pertenece a este usuario.", "danger")
        return redirect(url_for("main.subscription"))

    if has_pro_access(user):
        from .plans import current_plan_code, current_plan_label

        flash(
            gettext(
                "Tu plan %(plan)s está activo.",
                plan=current_plan_label(current_plan_code(user)),
            ),
            "success",
        )
    else:
        flash("Pago recibido. Estamos confirmando tu suscripción con Stripe.", "success")
    return redirect(url_for("main.dashboard"))


@main.route("/sales/<int:sale_id>/cancel", methods=["POST"])
@require_permission("cancel_sales")
def cancel_sale(sale_id):
    return _reverse_sale(
        sale_id,
        movement_type="SALE_CANCELLATION",
        success_message=gettext(
            "Venta cancelada. Stock devuelto al inventario."
        ),
    )


@main.route("/sales/<int:sale_id>/return", methods=["POST"])
@require_permission("process_returns")
def return_sale(sale_id):
    return _reverse_sale(
        sale_id,
        movement_type="RETURN",
        success_message=gettext(
            "Devolución registrada. Stock devuelto al inventario."
        ),
    )


def _reverse_sale(sale_id, *, movement_type, success_message):
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    access_block = _trial_access_response(user)
    if access_block:
        return access_block
    organization_id = current_organization_id(user)
    membership = active_membership(user)
    sale = Sale.query.filter_by(
        id=sale_id, organization_id=organization_id
    ).first_or_404()
    payment_method = (
        sale.sales_ticket.payment_method
        if sale.sales_ticket
        else sale.payment_method
    )
    cash_session = None
    if payment_method == "cash":
        cash_session = open_cash_session(organization_id, lock=True)
        if not cash_session:
            flash(
                gettext("Abre la caja antes de devolver una venta en efectivo."),
                "danger",
            )
            return redirect(url_for("cash.index"))
    if payment_method == "credit" and sale.sales_ticket:
        customer = sale.sales_ticket.customer
        if not customer:
            flash(
                gettext("La venta a crédito no tiene un cliente asociado."),
                "danger",
            )
            return redirect(url_for("main.sell"))
        try:
            record_credit_reversal(
                customer,
                membership,
                sale.total,
                sale.sales_ticket,
            )
        except CreditError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("main.sell"))
    product = (
        Product.query.filter_by(
            id=sale.product_id, organization_id=organization_id
        )
        .with_for_update()
        .first()
    )
    if product:
        stock_before = product.stock
        product.stock += sale.quantity
        record_inventory_movement(
            product,
            membership,
            movement_type,
            stock_before,
            product.stock,
            reason=(
                gettext("Devolución de venta")
                if movement_type == "RETURN"
                else gettext("Cancelación de venta")
            ),
            sales_ticket=sale.sales_ticket,
        )
    if payment_method == "cash":
        record_cash_movement(
            cash_session,
            membership,
            "REFUND",
            sale.total,
            note=gettext(
                "Devolución de %(folio)s",
                folio=_short_sale_folio([sale]),
            ),
            sales_ticket=sale.sales_ticket,
        )
    db.session.execute(
        update(InventoryMovement)
        .where(
            InventoryMovement.organization_id == organization_id,
            InventoryMovement.sale_id == sale.id,
        )
        .values(sale_id=None)
    )
    db.session.delete(sale)
    db.session.commit()
    flash(success_message, "success")
    return redirect(url_for("main.sell"))


@main.route("/subscription")
@require_permission("manage_subscription")
def subscription():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    user = current_organization_owner(user)
    subscription_info = None
    if user.stripe_subscription_id and not current_app.config["STRIPE_DISABLED"]:
        stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
        try:
            sub = stripe.Subscription.retrieve(user.stripe_subscription_id)
            _validate_subscription(sub, user=user)
            _sync_subscription_state(user, sub, datetime.utcnow())
            db.session.commit()
            subscription_info = sub
        except Exception:
            db.session.rollback()
            current_app.logger.exception("No se pudo sincronizar la suscripción")
            flash("No pudimos actualizar el estado de tu suscripción.", "danger")
    from .plans import (
        commercial_plans,
        current_plan_code,
        current_plan_label,
        has_entitlement,
        plan_price,
    )

    paid_access = has_pro_access(user)
    plan_code = current_plan_code(user, has_paid_access=paid_access)
    organization = Organization.query.filter_by(
        owner_user_id=user.id
    ).first()
    last_monthly_report = (
        MonthlyOwnerReport.query.filter_by(
            organization_id=organization.id
        )
        .order_by(
            MonthlyOwnerReport.report_year.desc(),
            MonthlyOwnerReport.report_month.desc(),
        )
        .first()
        if organization
        else None
    )
    return render_template(
        "subscription.html",
        user=user,
        subscription_info=subscription_info,
        has_paid_access=paid_access,
        current_plan_code=plan_code,
        current_plan_label=current_plan_label(plan_code),
        current_plan_price=plan_price(plan_code),
        commercial_plans=commercial_plans(current_app.config),
        organization=organization,
        monthly_report_available=(
            paid_access
            and has_entitlement(
                user,
                "monthly_owner_report",
                has_paid_access=paid_access,
            )
        ),
        last_monthly_report=last_monthly_report,
        subscription_status_label=_subscription_status_label(user.subscription_status),
        current_period_end_local=utc_to_local(
            user.current_period_end,
            user.timezone,
        ),
        next_payment_attempt_local=utc_to_local(
            user.next_payment_attempt,
            user.timezone,
        ),
    )


@main.post("/subscription/monthly-report")
@require_permission("manage_subscription")
def update_monthly_report_settings():
    from .plans import has_entitlement

    actor = current_user()
    owner = current_organization_owner(actor) if actor else None
    membership = active_membership(actor) if actor else None
    if not owner or not membership:
        return redirect(url_for("main.login"))
    enabled = request.form.get("enabled") == "1"
    paid_access = has_pro_access(owner)
    if enabled and (
        not paid_access
        or not has_entitlement(
            owner,
            "monthly_owner_report",
            has_paid_access=paid_access,
        )
    ):
        flash(
            "El reporte mensual para el propietario está incluido en PATIA Pro.",
            "info",
        )
        return redirect(url_for("main.subscribe"))
    recipient = (
        request.form.get("recipient", "").strip().lower() or owner.email
    )
    try:
        validate_email(recipient, check_deliverability=False)
    except EmailNotValidError:
        flash("Escribe un correo válido para recibir el reporte.", "danger")
        return redirect(url_for("main.subscription"))
    organization = membership.organization
    organization.monthly_report_enabled = enabled
    organization.monthly_report_recipient = (
        recipient if recipient != owner.email else None
    )
    db.session.commit()
    flash(
        "Preferencias del reporte mensual actualizadas.",
        "success",
    )
    return redirect(url_for("main.subscription"))


@main.post("/subscription/change-plan")
@require_permission("manage_subscription")
def change_subscription_plan():
    from .plans import (
        PAID_PLAN_CODES,
        STARTER,
        current_plan_code,
        entitlements_for,
        price_id_for,
    )

    actor = current_user()
    owner = current_organization_owner(actor) if actor else None
    membership = active_membership(actor) if actor else None
    target_plan = str(request.form.get("plan_code") or "").upper()
    if not owner or not membership or target_plan not in PAID_PLAN_CODES:
        flash("Selecciona un plan válido.", "danger")
        return redirect(url_for("main.subscription"))
    if not owner.stripe_subscription_id:
        return redirect(url_for("main.subscribe"))
    if current_app.config["STRIPE_DISABLED"]:
        flash("La facturación no está disponible en este entorno.", "danger")
        return redirect(url_for("main.subscription"))
    target_price = price_id_for(current_app.config, target_plan)
    if not target_price:
        flash(
            "Este plan estará disponible para contratación muy pronto.",
            "info",
        )
        return redirect(url_for("main.subscription"))
    current_code = current_plan_code(
        owner, has_paid_access=has_pro_access(owner)
    )
    if current_code == target_plan and not owner.pending_plan_code:
        flash("Ese ya es tu plan actual.", "info")
        return redirect(url_for("main.subscription"))

    if target_plan == STARTER:
        member_count = OrganizationMember.query.filter_by(
            organization_id=membership.organization_id,
            is_active=True,
        ).count()
        if member_count > entitlements_for(STARTER).max_members:
            flash(
                "Tu negocio tiene más personas de las permitidas en Starter. "
                "Desactiva las necesarias antes de solicitar el cambio.",
                "warning",
            )
            return redirect(url_for("team.index"))
        active_manager = OrganizationMember.query.filter_by(
            organization_id=membership.organization_id,
            is_active=True,
            role="MANAGER",
        ).first()
        if active_manager:
            flash(
                gettext(
                    "Starter no incluye encargados. Cambia sus accesos a Cajero antes de solicitar el cambio."
                ),
                "warning",
            )
            return redirect(url_for("team.index"))

    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    try:
        subscription = stripe.Subscription.retrieve(
            owner.stripe_subscription_id
        )
        _validate_subscription(subscription, user=owner)
        items = (subscription.get("items") or {}).get("data") or []
        if len(items) != 1 or not items[0].get("id"):
            raise StripeEventIgnored(
                "La suscripción no tiene un único concepto modificable."
            )
        metadata = {
            **(subscription.get("metadata") or {}),
            "user_id": str(owner.id),
            "organization_id": str(membership.organization_id),
            "plan_code": target_plan,
        }
        period_end = _subscription_period_end(subscription)
        if target_plan == STARTER:
            schedule_id = subscription.get("schedule")
            if schedule_id:
                schedule = stripe.SubscriptionSchedule.retrieve(
                    schedule_id
                )
            else:
                schedule = stripe.SubscriptionSchedule.create(
                    from_subscription=owner.stripe_subscription_id,
                    idempotency_key=(
                        f"patia-plan-schedule-{owner.id}-starter-"
                        f"{subscription.get('current_period_end')}"
                    ),
                )
            current_price = (items[0].get("price") or {}).get("id")
            stripe.SubscriptionSchedule.modify(
                schedule.get("id"),
                end_behavior="release",
                phases=[
                    {
                        "start_date": subscription.get(
                            "current_period_start"
                        ),
                        "end_date": subscription.get("current_period_end"),
                        "items": [
                            {
                                "price": current_price,
                                "quantity": items[0].get("quantity", 1),
                            }
                        ],
                        "metadata": subscription.get("metadata") or {},
                    },
                    {
                        "start_date": subscription.get(
                            "current_period_end"
                        ),
                        "items": [
                            {
                                "price": target_price,
                                "quantity": items[0].get("quantity", 1),
                            }
                        ],
                        "metadata": metadata,
                    },
                ],
                idempotency_key=(
                    f"patia-plan-downgrade-{owner.id}-starter-"
                    f"{subscription.get('current_period_end')}"
                ),
            )
            owner.pending_plan_code = target_plan
            owner.pending_plan_effective_at = period_end
            flash(
                "El cambio a Starter quedó programado para el final de tu periodo pagado.",
                "success",
            )
        else:
            stripe.Subscription.modify(
                owner.stripe_subscription_id,
                items=[{"id": items[0]["id"], "price": target_price}],
                metadata=metadata,
                proration_behavior="create_prorations",
                payment_behavior="pending_if_incomplete",
                idempotency_key=(
                    f"patia-plan-upgrade-{owner.id}-pro-"
                    f"{subscription.get('current_period_end')}"
                ),
            )
            owner.pending_plan_code = target_plan
            owner.pending_plan_effective_at = datetime.utcnow()
            flash(
                "Solicitamos el cambio a Pro. Se activará cuando Stripe confirme el pago.",
                "success",
            )
        db.session.commit()
    except Exception as error:
        db.session.rollback()
        error_type, error_code, error_message = _safe_stripe_error(error)
        current_app.logger.exception(
            "No se pudo cambiar el plan: type=%s code=%s message=%s",
            error_type,
            error_code,
            error_message,
        )
        flash(
            "No pudimos cambiar el plan. Intenta nuevamente.",
            "danger",
        )
    return redirect(url_for("main.subscription"))


@main.route("/cancel-subscription", methods=["POST"])
@require_permission("manage_subscription")
def cancel_subscription():
    user = current_user()
    user = current_organization_owner(user) if user else None
    if not user or not user.stripe_subscription_id:
        return redirect(url_for("main.dashboard"))
    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    try:
        stripe.Subscription.modify(user.stripe_subscription_id, cancel_at_period_end=True)
        user.cancel_at_period_end = True
        db.session.commit()
        flash("Tu suscripcion se cancelara al final del periodo pagado.", "success")
    except Exception:
        db.session.rollback()
        current_app.logger.exception("No se pudo programar la cancelación")
        flash("No pudimos cancelar la suscripción. Intenta nuevamente.", "danger")
    return redirect(url_for("main.subscription"))


@main.route("/reactivate-subscription", methods=["POST"])
@require_permission("manage_subscription")
def reactivate_subscription():
    user = current_user()
    user = current_organization_owner(user) if user else None
    if not user or not user.stripe_subscription_id:
        return redirect(url_for("main.dashboard"))
    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    try:
        stripe.Subscription.modify(user.stripe_subscription_id, cancel_at_period_end=False)
        user.cancel_at_period_end = False
        db.session.commit()
        flash("Tu suscripcion ha sido reactivada.", "success")
    except Exception:
        db.session.rollback()
        current_app.logger.exception("No se pudo reactivar la suscripción")
        flash("No pudimos reactivar la suscripción. Intenta nuevamente.", "danger")
    return redirect(url_for("main.subscription"))


@main.route("/billing-portal", methods=["POST"])
@require_permission("manage_subscription")
def billing_portal():
    user = current_user()
    user = current_organization_owner(user) if user else None
    if not user or not user.stripe_customer_id:
        return redirect(url_for("main.subscribe"))
    return _redirect_to_billing_portal(user)


def _redirect_to_billing_portal(user):
    if current_app.config["STRIPE_DISABLED"]:
        flash("La facturación no está disponible en este entorno.", "danger")
        return redirect(url_for("main.subscription"))
    stripe.api_key = current_app.config["STRIPE_SECRET_KEY"]
    try:
        portal = stripe.billing_portal.Session.create(
            customer=user.stripe_customer_id,
            return_url=_public_url("/subscription"),
        )
        return redirect(portal.url, code=303)
    except Exception as error:
        error_type, error_code, error_message = _safe_stripe_error(error)
        current_app.logger.exception(
            "No se pudo abrir el portal de facturación: type=%s code=%s message=%s",
            error_type,
            error_code,
            error_message,
        )
        flash("No pudimos abrir el portal de facturación.", "danger")
        return redirect(url_for("main.subscription"))


@main.route("/admin")
def admin():
    user = current_user()
    if not user:
        session.clear()
        return redirect(url_for("main.login"))
    if user.email != "albertonicopat@gmail.com":
        flash("No autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    organizations = (
        Organization.query.options(selectinload(Organization.owner))
        .order_by(Organization.created_at.desc())
        .all()
    )
    product_counts = dict(
        db.session.query(Product.organization_id, func.count(Product.id))
        .filter(Product.is_active.is_(True))
        .group_by(Product.organization_id)
        .all()
    )
    sale_totals = {
        organization_id: (count, total or 0)
        for organization_id, count, total in (
            db.session.query(
                Sale.organization_id,
                func.count(Sale.id),
                func.sum(Sale.total),
            )
            .join(Organization, Organization.id == Sale.organization_id)
            .filter(Sale.currency_code == Organization.currency_code)
            .group_by(Sale.organization_id)
            .all()
        )
    }
    member_counts = dict(
        db.session.query(
            OrganizationMember.organization_id,
            func.count(OrganizationMember.id),
        )
        .filter(OrganizationMember.is_active.is_(True))
        .group_by(OrganizationMember.organization_id)
        .all()
    )
    customer_counts = dict(
        db.session.query(Customer.organization_id, func.count(Customer.id))
        .filter(Customer.is_active.is_(True))
        .group_by(Customer.organization_id)
        .all()
    )
    last_activity = dict(
        db.session.query(
            Sale.organization_id, func.max(Sale.created_at)
        )
        .group_by(Sale.organization_id)
        .all()
    )
    latest_report_ids = (
        db.session.query(
            MonthlyOwnerReport.organization_id,
            func.max(MonthlyOwnerReport.id).label("report_id"),
        )
        .group_by(MonthlyOwnerReport.organization_id)
        .subquery()
    )
    latest_reports = {
        report.organization_id: report
        for report in (
            MonthlyOwnerReport.query.join(
                latest_report_ids,
                MonthlyOwnerReport.id == latest_report_ids.c.report_id,
            ).all()
        )
    }
    today = datetime.utcnow()
    next_report_estimate = (
        (today.replace(day=28) + timedelta(days=4))
        .replace(day=2, hour=13, minute=0, second=0, microsecond=0)
    )
    from .plans import (
        PRO,
        STARTER,
        current_plan_code,
        current_plan_label,
        plan_price,
    )

    clients = []
    total_products = total_sales_count = total_sales_money = trial_clients = expired_clients = expiring_soon = new_this_week = new_this_month = 0

    for organization in organizations:
        u = organization.owner
        products_count = product_counts.get(organization.id, 0)
        sales_count, sales_money = sale_totals.get(organization.id, (0, 0))
        days_in_patia = (today - u.created_at).days if u.created_at else 0
        trial_days_left = max(0, 14 - days_in_patia)

        paid_access = has_pro_access(u)
        if paid_access:
            status = "Activo"
            trial_days_left = "inf"
        elif trial_days_left > 0:
            status = "Prueba"
            trial_clients += 1
        else:
            status = "Vencido"
            expired_clients += 1

        if trial_days_left != "inf" and 0 < trial_days_left <= 7:
            expiring_soon += 1
        if days_in_patia <= 7:
            new_this_week += 1
        if days_in_patia <= 30:
            new_this_month += 1

        total_products += products_count
        total_sales_count += sales_count
        total_sales_money += sales_money
        plan_code = current_plan_code(u, has_paid_access=paid_access)
        monthly_report = latest_reports.get(organization.id)
        attention_reasons = []
        if u.subscription_status == "past_due":
            attention_reasons.append(("payment", gettext("Pago pendiente")))
        if trial_days_left != "inf" and 0 < trial_days_left <= 3:
            attention_reasons.append(
                (
                    "trial",
                    gettext(
                        "La prueba vence en %(days)s días",
                        days=trial_days_left,
                    ),
                )
            )
        if u.cancel_at_period_end:
            attention_reasons.append(
                ("cancellation", gettext("Cancelación programada"))
            )
        if monthly_report and monthly_report.status == "failed":
            attention_reasons.append(
                ("report", gettext("Reporte mensual fallido"))
            )
        if status == "Vencido":
            attention_reasons.append(("expired", gettext("Cuenta vencida")))

        plan_admin_labels = {
            "TRIAL": gettext("Trial"),
            "STARTER": gettext("Starter"),
            "PRO": gettext("Pro"),
            "GRANDFATHERED": gettext("Cliente anterior"),
            "MANUAL": gettext("Acceso manual"),
        }
        status_code = (
            "payment_pending"
            if u.subscription_status == "past_due"
            else "cancelling"
            if u.cancel_at_period_end
            else "expired"
            if status == "Vencido"
            else "trial"
            if plan_code == "TRIAL"
            else "active"
        )
        clients.append({
            "organization": organization,
            "user": u,
            "products_count": products_count,
            "sales_count": sales_count,
            "sales_money": sales_money,
            "days_in_patia": days_in_patia,
            "trial_days_left": trial_days_left,
            "trial_end": (
                u.created_at + timedelta(days=14) if u.created_at else None
            ),
            "status": status,
            "plan_code": plan_code,
            "plan_label": plan_admin_labels.get(
                plan_code, current_plan_label(plan_code)
            ),
            "subscription_status": u.subscription_status,
            "renewal_date": utc_to_local(
                u.current_period_end, u.timezone
            ),
            "scheduled_cancellation": u.cancel_at_period_end,
            "in_grace": u.subscription_status == "past_due",
            "manual_access": u.manual_pro_access,
            "read_only": status == "Vencido",
            "price": plan_price(plan_code),
            "member_count": member_counts.get(organization.id, 0),
            "customer_count": customer_counts.get(organization.id, 0),
            "last_activity": last_activity.get(organization.id),
            "last_payment": u.stripe_invoice_updated_at,
            "monthly_report_enabled": organization.monthly_report_enabled,
            "monthly_report": monthly_report,
            "next_monthly_report_at": (
                next_report_estimate
                if organization.monthly_report_enabled
                and plan_code == PRO
                and not status == "Vencido"
                else None
            ),
            "status_code": status_code,
            "status_label": {
                "payment_pending": gettext("Pago pendiente"),
                "cancelling": gettext("Cancelación programada"),
                "expired": gettext("Vencido"),
                "trial": gettext("En prueba"),
                "active": gettext("Activo"),
            }[status_code],
            "attention_reasons": attention_reasons,
        })

    top_client = max(clients, key=lambda c: c["products_count"], default=None)
    latest_client = clients[0] if clients else None

    mrr = sum(
        client["price"] or 0
        for client in clients
        if client["plan_code"] in {STARTER, PRO}
        and client["subscription_status"] in {"active", "trialing"}
        and not client["read_only"]
    )
    paying_clients = sum(
        client["plan_code"] in {STARTER, PRO}
        and client["subscription_status"] in {"active", "trialing"}
        and not client["read_only"]
        for client in clients
    )
    plan_counts = {
        code: sum(
            client["plan_code"] == code and not client["read_only"]
            for client in clients
        )
        for code in ("TRIAL", "STARTER", "PRO", "GRANDFATHERED", "MANUAL")
    }
    renewal_limit = today + timedelta(days=30)
    upcoming_renewals = sum(
        bool(
            client["user"].current_period_end
            and today <= client["user"].current_period_end <= renewal_limit
            and client["subscription_status"]
            in {"active", "trialing", "past_due"}
        )
        for client in clients
    )
    report_sent = sum(
        bool(client["monthly_report"] and client["monthly_report"].status == "sent")
        for client in clients
    )
    report_failed = sum(
        bool(client["monthly_report"] and client["monthly_report"].status == "failed")
        for client in clients
    )
    attention_clients = sorted(
        (client for client in clients if client["attention_reasons"]),
        key=lambda client: (
            0
            if client["status_code"] == "payment_pending"
            else 1
            if client["status_code"] == "expired"
            else 2,
            client["organization"].name.lower(),
        ),
    )
    query = request.args.get("q", "").strip()
    selected_plan = request.args.get("plan", "all").strip().upper()
    selected_status = request.args.get("status", "all").strip().lower()
    selected_attention = request.args.get("attention", "all").strip().lower()
    valid_plans = {"ALL", "TRIAL", "STARTER", "PRO", "GRANDFATHERED", "MANUAL"}
    valid_statuses = {
        "all",
        "active",
        "payment_pending",
        "cancelling",
        "expired",
        "trial",
    }
    valid_attention = {
        "all",
        "payment",
        "trial",
        "cancellation",
        "report",
        "expired",
    }
    if selected_plan not in valid_plans:
        selected_plan = "ALL"
    if selected_status not in valid_statuses:
        selected_status = "all"
    if selected_attention not in valid_attention:
        selected_attention = "all"

    visible_clients = clients
    if query:
        normalized_query = query.casefold()
        visible_clients = [
            client
            for client in visible_clients
            if normalized_query
            in " ".join(
                (
                    client["organization"].name,
                    client["user"].email,
                    client["user"].first_name or "",
                    client["user"].last_name or "",
                    client["user"].phone or "",
                )
            ).casefold()
        ]
    if selected_plan != "ALL":
        visible_clients = [
            client
            for client in visible_clients
            if client["plan_code"] == selected_plan
        ]
    if selected_status != "all":
        visible_clients = [
            client
            for client in visible_clients
            if client["status_code"] == selected_status
        ]
    if selected_attention != "all":
        visible_clients = [
            client
            for client in visible_clients
            if any(
                reason_code == selected_attention
                for reason_code, _ in client["attention_reasons"]
            )
        ]

    applied_filters = []
    if query:
        applied_filters.append(("q", query, gettext("Búsqueda: %(value)s", value=query)))
    if selected_plan != "ALL":
        applied_filters.append(
            (
                "plan",
                selected_plan,
                gettext(
                    "Plan: %(value)s",
                    value=next(
                        (
                            client["plan_label"]
                            for client in clients
                            if client["plan_code"] == selected_plan
                        ),
                        selected_plan.title(),
                    ),
                ),
            )
        )
    if selected_status != "all":
        status_labels = {
            "active": gettext("Activo"),
            "payment_pending": gettext("Pago pendiente"),
            "cancelling": gettext("Cancelación programada"),
            "expired": gettext("Vencido"),
            "trial": gettext("En prueba"),
        }
        applied_filters.append(
            (
                "status",
                selected_status,
                gettext(
                    "Estado: %(value)s",
                    value=status_labels[selected_status],
                ),
            )
        )
    if selected_attention != "all":
        attention_labels = {
            "payment": gettext("Pago pendiente"),
            "trial": gettext("Prueba por vencer"),
            "cancellation": gettext("Cancelación programada"),
            "report": gettext("Reporte fallido"),
            "expired": gettext("Cuenta vencida"),
        }
        applied_filters.append(
            (
                "attention",
                selected_attention,
                gettext(
                    "Atención: %(value)s",
                    value=attention_labels[selected_attention],
                ),
            )
        )

    return render_template("admin.html", clients=visible_clients, total_clients=len(organizations), total_products=total_products,
        total_sales_count=total_sales_count, total_sales_money=total_sales_money, trial_clients=trial_clients,
        expired_clients=expired_clients, expiring_soon=expiring_soon, new_this_week=new_this_week,
        new_this_month=new_this_month, top_client=top_client, latest_client=latest_client,
        plan_counts=plan_counts, estimated_mrr=mrr, report_sent=report_sent,
        report_failed=report_failed,
        upcoming_renewals=upcoming_renewals,
        scheduled_cancellations=sum(c["scheduled_cancellation"] for c in clients),
        past_due_count=sum(c["subscription_status"] == "past_due" for c in clients),
        paying_clients=paying_clients,
        attention_clients=attention_clients,
        query=query,
        selected_plan=selected_plan,
        selected_status=selected_status,
        selected_attention=selected_attention,
        applied_filters=applied_filters)


@main.route("/admin/organizations/<int:organization_id>")
def admin_organization_detail(organization_id):
    admin_user = current_user()
    if not admin_user:
        session.clear()
        return redirect(url_for("main.login"))
    if admin_user.email != "albertonicopat@gmail.com":
        flash("No autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    organization = (
        Organization.query.options(
            selectinload(Organization.owner),
            selectinload(Organization.members),
        )
        .filter_by(id=organization_id)
        .first_or_404()
    )
    owner = organization.owner
    usage = {
        "members": sum(member.is_active for member in organization.members),
        "products": Product.query.filter_by(
            organization_id=organization.id, is_active=True
        ).count(),
        "sales": Sale.query.filter_by(
            organization_id=organization.id
        ).count(),
        "customers": Customer.query.filter_by(
            organization_id=organization.id, is_active=True
        ).count(),
    }
    last_sale = (
        Sale.query.filter_by(organization_id=organization.id)
        .order_by(Sale.created_at.desc(), Sale.id.desc())
        .first()
    )
    report_history = (
        MonthlyOwnerReport.query.filter_by(
            organization_id=organization.id
        )
        .order_by(
            MonthlyOwnerReport.report_year.desc(),
            MonthlyOwnerReport.report_month.desc(),
        )
        .limit(12)
        .all()
    )
    from .plans import (
        current_plan_code,
        plan_price,
        subscription_access_is_active,
    )

    paid_access = subscription_access_is_active(
        owner,
        grace_days=current_app.config.get("STRIPE_PAST_DUE_GRACE_DAYS", 3),
    )
    plan_code = current_plan_code(owner, has_paid_access=paid_access)
    plan_label = {
        "TRIAL": gettext("Trial"),
        "STARTER": gettext("Starter"),
        "PRO": gettext("Pro"),
        "GRANDFATHERED": gettext("Cliente anterior"),
        "MANUAL": gettext("Acceso manual"),
    }.get(plan_code, gettext("Plan actual"))
    return render_template(
        "admin_organization.html",
        organization=organization,
        owner=owner,
        usage=usage,
        last_sale=last_sale,
        report_history=report_history,
        plan_code=plan_code,
        plan_label=plan_label,
        plan_price=plan_price(plan_code),
        paid_access=paid_access,
    )


@main.route(
    "/admin/monthly-reports/<int:report_id>/retry",
    methods=["POST"],
)
def admin_retry_monthly_report(report_id):
    admin_user = current_user()
    if not admin_user or admin_user.email != "albertonicopat@gmail.com":
        return redirect(url_for("main.dashboard"))
    report = db.session.get(MonthlyOwnerReport, report_id)
    if not report:
        abort(404)
    if report.status != "failed":
        flash("Este reporte no necesita reintentarse.", "warning")
        return redirect(
            url_for(
                "main.admin_organization_detail",
                organization_id=report.organization_id,
            )
        )
    organization_id = report.organization_id
    try:
        from .monthly_reports import generate_monthly_report

        retried, _ = generate_monthly_report(
            organization_id,
            report.report_year,
            report.report_month,
            send=True,
            force_retry=True,
        )
        if retried.status == "sent":
            flash("Reporte mensual enviado correctamente.", "success")
        else:
            flash(
                "No pudimos entregar el reporte. Podrás reintentarlo después.",
                "danger",
            )
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "Admin monthly report retry failed for report_id=%s",
            report_id,
        )
        flash(
            "No pudimos entregar el reporte. Podrás reintentarlo después.",
            "danger",
        )
    return redirect(
        url_for(
            "main.admin_organization_detail",
            organization_id=organization_id,
        )
    )


@main.route("/products/<int:product_id>/delete", methods=["POST"])
@require_permission("manage_inventory")
def delete_product(product_id):
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    organization_id = current_organization_id(user)
    access_block = _trial_access_response(user)
    if access_block:
        return access_block
    product = Product.query.filter_by(
        id=product_id,
        organization_id=organization_id,
        is_active=True,
    ).first_or_404()
    if Sale.query.filter_by(product_id=product.id, organization_id=organization_id).first():
        product.is_active = False
        db.session.commit()
        flash("Producto retirado del catálogo. Su historial de ventas se conserva.", "success")
        return redirect(url_for("main.products") + "#catalogo")
    db.session.execute(
        update(InventoryMovement)
        .where(
            InventoryMovement.organization_id == organization_id,
            InventoryMovement.product_id == product.id,
        )
        .values(product_id=None)
    )
    db.session.delete(product)
    db.session.commit()
    flash("Producto eliminado correctamente.", "success")
    return redirect(url_for("main.products") + "#catalogo")


@main.route("/products/delete-all", methods=["POST"])
@require_permission("manage_inventory")
def delete_all_products():
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    organization_id = current_organization_id(user)
    access_block = _trial_access_response(user)
    if access_block:
        return access_block
    products = Product.query.filter_by(organization_id=organization_id, is_active=True).all()
    product_ids_with_sales = {
        product_id
        for (product_id,) in (
            db.session.query(Sale.product_id)
            .filter(
                Sale.organization_id == organization_id,
                Sale.product_id.in_([product.id for product in products]),
            )
            .distinct()
            .all()
        )
    }
    deleted = 0
    protected = 0
    deleted_ids = []
    for product in products:
        if product.id in product_ids_with_sales:
            product.is_active = False
            protected += 1
            continue
        db.session.delete(product)
        deleted_ids.append(product.id)
        deleted += 1
    if deleted_ids:
        db.session.execute(
            update(InventoryMovement)
            .where(
                InventoryMovement.organization_id == organization_id,
                InventoryMovement.product_id.in_(deleted_ids),
            )
            .values(product_id=None)
        )
    db.session.commit()
    if protected:
        flash(
            gettext(
                "Eliminamos %(deleted)s productos sin ventas y retiramos %(protected)s del catálogo conservando su historial.",
                deleted=deleted,
                protected=protected,
            ),
            "info",
        )
    else:
        flash(
            gettext("Eliminamos %(count)s productos del catálogo.", count=deleted),
            "success",
        )
    return redirect(url_for("main.products"))


@main.route("/products/delete-selected", methods=["POST"])
@require_permission("manage_inventory")
def delete_selected_products():
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    organization_id = current_organization_id(user)
    access_block = _trial_access_response(user)
    if access_block:
        return access_block
    ids_raw = request.form.get("ids", "")
    ids = [int(i) for i in ids_raw.split(",") if i.strip().isdigit()]
    products = (
        Product.query.filter(
            Product.id.in_(ids),
            Product.organization_id == organization_id,
            Product.is_active.is_(True),
        ).all()
        if ids
        else []
    )
    product_ids_with_sales = {
        product_id
        for (product_id,) in (
            db.session.query(Sale.product_id)
            .filter(
                Sale.organization_id == organization_id,
                Sale.product_id.in_([product.id for product in products]),
            )
            .distinct()
            .all()
        )
    }
    deleted = 0
    protected = 0
    deleted_ids = []
    for product in products:
        if product.id in product_ids_with_sales:
            product.is_active = False
            protected += 1
            continue
        db.session.delete(product)
        deleted_ids.append(product.id)
        deleted += 1
    if deleted_ids:
        db.session.execute(
            update(InventoryMovement)
            .where(
                InventoryMovement.organization_id == organization_id,
                InventoryMovement.product_id.in_(deleted_ids),
            )
            .values(product_id=None)
        )
    db.session.commit()
    if protected:
        flash(
            gettext(
                "Eliminamos %(deleted)s seleccionados y retiramos %(protected)s conservando su historial.",
                deleted=deleted,
                protected=protected,
            ),
            "info",
        )
    else:
        flash(gettext("%(count)s productos eliminados.", count=deleted), "success")
    return redirect(url_for("main.products"))


@main.route("/suppliers/<int:supplier_id>/delete", methods=["POST"])
@require_permission("manage_inventory")
def delete_supplier(supplier_id):
    if not session.get("user_id"):
        return redirect(url_for("main.login"))
    user = current_user()
    organization_id = current_organization_id(user)
    access_block = _trial_access_response(user)
    if access_block:
        return access_block
    supplier = Supplier.query.filter_by(
        id=supplier_id, organization_id=organization_id
    ).first_or_404()
    db.session.delete(supplier)
    db.session.commit()
    flash("Proveedor eliminado correctamente.", "success")
    return redirect(url_for("main.suppliers"))


@main.route("/admin/delete-user/<int:user_id>", methods=["POST"])
def admin_delete_user(user_id):
    admin_user = current_user()
    if not admin_user or admin_user.email != "albertonicopat@gmail.com":
        return redirect(url_for("main.dashboard"))
    user = User.query.get_or_404(user_id)
    if user.email == "albertonicopat@gmail.com":
        flash("No puedes eliminar tu cuenta de administrador.")
        return redirect(url_for("main.admin"))
    if _has_managed_stripe_subscription(user):
        flash(
            "No puedes eliminar este usuario mientras tenga una suscripción Stripe gestionable.",
            "danger",
        )
        return redirect(url_for("main.admin"))
    organization = Organization.query.filter_by(owner_user_id=user.id).first()
    if organization:
        flash(
            "No puedes eliminar al propietario mientras su organizaciÃ³n conserve datos o integrantes.",
            "danger",
        )
        return redirect(url_for("main.admin"))
    db.session.delete(user)
    db.session.commit()
    flash("Cliente eliminado correctamente.")
    return redirect(url_for("main.admin"))


@main.route("/admin/make-pro/<int:user_id>", methods=["POST"])
def admin_make_pro(user_id):
    admin_user = current_user()
    if not admin_user or admin_user.email != "albertonicopat@gmail.com":
        return redirect(url_for("main.dashboard"))
    user = User.query.get_or_404(user_id)
    organization = Organization.query.filter_by(owner_user_id=user.id).first()
    if not organization:
        flash(
            gettext(
                "El acceso Pro se administra en el propietario de la organización."
            ),
            "danger",
        )
        return redirect(url_for("main.admin"))
    user.manual_pro_access = True
    sync_user_plan(user)
    db.session.commit()
    flash(gettext("Acceso manual Pro activado."), "success")
    return redirect(
        url_for(
            "main.admin_organization_detail",
            organization_id=organization.id,
        )
    )


@main.route("/admin/remove-manual-pro/<int:user_id>", methods=["POST"])
def admin_remove_manual_pro(user_id):
    admin_user = current_user()
    if not admin_user or admin_user.email != "albertonicopat@gmail.com":
        return redirect(url_for("main.dashboard"))
    user = User.query.get_or_404(user_id)
    organization = Organization.query.filter_by(owner_user_id=user.id).first()
    if not organization:
        flash(
            gettext(
                "El acceso Pro se administra en el propietario de la organización."
            ),
            "danger",
        )
        return redirect(url_for("main.admin"))
    user.manual_pro_access = False
    sync_user_plan(user)
    db.session.commit()
    flash(gettext("Acceso manual Pro desactivado."), "success")
    return redirect(
        url_for(
            "main.admin_organization_detail",
            organization_id=organization.id,
        )
    )

@main.route("/settings", methods=["GET", "POST"])
@require_roles("OWNER")
def settings():
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    membership = active_membership(user)
    user = membership.organization.owner
    if request.method == "POST":
        access_block = _trial_access_response(user)
        if access_block:
            return access_block
        company_name = request.form.get("company_name", "").strip()
        if not company_name:
            flash("El nombre del negocio es obligatorio.", "danger")
            return redirect(url_for("main.settings"))
        user.company_name = company_name
        user.rfc = request.form.get("rfc", "").strip().upper()
        user.tax_regime = request.form.get("tax_regime", "").strip()
        user.address = request.form.get("address", "").strip()
        user.city = request.form.get("city", "").strip()
        user.state = request.form.get("state", "").strip()
        user.postal_code = request.form.get("postal_code", "").strip()
        user.phone = request.form.get("phone", "").strip()
        user.timezone = safe_timezone_name(request.form.get("timezone"))
        country_code = str(
            request.form.get("country_code")
            or membership.organization.country_code
            or "MX"
        ).upper()
        currency_code = str(
            request.form.get("currency_code")
            or membership.organization.currency_code
            or membership.organization.currency
            or "MXN"
        ).upper()
        if country_code not in COUNTRY_OPTIONS or currency_code not in SUPPORTED_CURRENCIES:
            flash(
                gettext("Selecciona un país y una moneda compatibles."),
                "danger",
            )
            return redirect(url_for("main.settings"))
        _, suggested_locale = country_defaults(country_code)
        locale_code = normalize_locale_code(
            request.form.get("locale_code") or suggested_locale,
            currency_code,
        )
        membership.organization.name = company_name
        membership.organization.timezone = user.timezone
        membership.organization.country_code = country_code
        membership.organization.currency_code = currency_code
        membership.organization.locale_code = locale_code
        membership.organization.currency = currency_code
        db.session.commit()
        flash("Configuración guardada.", "success")
        return redirect(url_for("main.settings"))
    return render_template(
        "settings.html",
        user=user,
        organization=membership.organization,
        country_options=COUNTRY_OPTIONS,
        supported_currencies=sorted(SUPPORTED_CURRENCIES),
        timezone_choices=_translated_timezone_choices(),
    )


@main.route("/receipt/<int:sale_id>")
@require_permission("use_pos")
def receipt(sale_id):
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))

    organization_id = current_organization_id(user)
    sale = Sale.query.filter_by(
        id=sale_id, organization_id=organization_id
    ).first_or_404()
    return redirect(url_for("main.ticket", ticket_ref=_sale_ticket_key(sale)))


@main.route("/ticket/<ticket_ref>")
@require_permission("use_pos")
def ticket(ticket_ref):
    user = current_user()
    if not user:
        return redirect(url_for("main.login"))
    organization_id = current_organization_id(user)
    owner = current_organization_owner(user)

    if ticket_ref.startswith("sale-") and ticket_ref[5:].isdigit():
        sale = (
            Sale.query.options(
                selectinload(Sale.product),
                selectinload(Sale.sales_ticket),
            )
            .filter_by(id=int(ticket_ref[5:]), organization_id=organization_id)
            .first_or_404()
        )
        if sale.sales_ticket_id:
            sales = (
                Sale.query.options(
                    selectinload(Sale.product),
                    selectinload(Sale.sales_ticket),
                )
                .filter_by(organization_id=organization_id, sales_ticket_id=sale.sales_ticket_id)
                .order_by(Sale.id)
                .all()
            )
        else:
            sales = [sale]
    else:
        ticket_header = SalesTicket.query.filter_by(
            organization_id=organization_id,
            public_id=ticket_ref,
        ).first()
        sales = (
            Sale.query.options(
                selectinload(Sale.product),
                selectinload(Sale.sales_ticket),
            )
            .filter(
                Sale.organization_id == organization_id,
                (
                    Sale.sales_ticket_id == ticket_header.id
                    if ticket_header
                    else Sale.ticket_id == ticket_ref
                ),
            )
            .order_by(Sale.id)
            .all()
        )
        if not sales:
            abort(404)

    address_parts = [
        cleaned
        for cleaned in (
            _ticket_business_value(owner.address),
            _ticket_business_value(owner.city),
            _ticket_business_value(owner.state),
            _ticket_business_value(owner.postal_code),
        )
        if cleaned
    ]
    return render_template(
        "ticket.html",
        user=owner,
        sales=sales,
        ticket_id=_sale_ticket_key(sales[0]),
        folio=_short_sale_folio(sales),
        ticket_total=money_sum(sale.total for sale in sales),
        ticket_subtotal=money_sum(sale.total for sale in sales),
        item_count=sum(sale.quantity for sale in sales),
        ticket_created_at=utc_to_local(
            min(sale.created_at for sale in sales),
            active_membership(user).organization.timezone,
        ),
        payment_method=_payment_method_label(
            sales[0].sales_ticket.payment_method
            if sales[0].sales_ticket
            else sales[0].payment_method
        ),
        ticket_customer=(
            sales[0].sales_ticket.customer
            if sales[0].sales_ticket
            else None
        ),
        ticket_cashier_name=(
            " ".join(
                value
                for value in (
                    sales[0].sales_ticket.cashier_member.user.first_name,
                    sales[0].sales_ticket.cashier_member.user.last_name,
                )
                if value
            )
            if (
                sales[0].sales_ticket
                and sales[0].sales_ticket.cashier_member
                and sales[0].sales_ticket.cashier_member.user
            )
            else None
        ),
        amount_received=(
            sales[0].sales_ticket.amount_received
            if sales[0].sales_ticket
            else None
        ),
        change_amount=(
            sales[0].sales_ticket.change_amount
            if sales[0].sales_ticket
            else None
        ),
        business_address=", ".join(address_parts),
        business_phone=_ticket_business_value(owner.phone),
        auto_print=request.args.get("print") == "1",
        ticket_currency_code=(
            sales[0].sales_ticket.currency_code
            if sales[0].sales_ticket else sales[0].currency_code
        ),
        ticket_locale_code=(
            sales[0].sales_ticket.locale_code
            if sales[0].sales_ticket else sales[0].locale_code
        ),
    )
