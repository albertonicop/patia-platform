import io
import json
import os
import re
import time
import unittest
from openpyxl import load_workbook


os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("SECRET_KEY", "inventory-tests-secret")
os.environ.setdefault("STRIPE_SECRET_KEY", "sk_test_patia")
os.environ.setdefault("STRIPE_PRICE_ID", "price_patia_pro")
os.environ.setdefault("STRIPE_WEBHOOK_SECRET", "whsec_patia")
os.environ.setdefault("PUBLIC_BASE_URL", "https://patia.test")

from app import create_app, db
from app.models import InventoryMovement, Product, Sale, User
from app.team.services import ensure_owner_organization


CSV_HEADER = (
    "SKU,Codigo de barras,Nombre del producto,Categoria,Proveedor,"
    "Costo,Precio de venta,Stock inicial,Stock minimo\n"
)


class InventoryImportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config.update(
            TESTING=True,
            WTF_CSRF_ENABLED=False,
            RATELIMIT_ENABLED=False,
        )
        cls.context = cls.app.app_context()
        cls.context.push()
        db.create_all()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.engine.dispose()
        cls.context.pop()

    def setUp(self):
        db.session.rollback()
        InventoryMovement.query.delete()
        Sale.query.delete()
        Product.query.delete()
        User.query.delete()
        db.session.commit()
        self.client = self.app.test_client()
        self.user = User(
            email="inventory@patia.test",
            company_name="Tienda Inventario",
            phone="5555555555",
            city="Puebla",
            state="Puebla",
            email_verified=True,
        )
        self.user.set_password("Password123")
        db.session.add(self.user)
        db.session.flush()
        ensure_owner_organization(self.user)
        db.session.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user.id

    def add_product(self, sku="EXISTE", barcode="7501", stock=5):
        product = Product(
            organization_id=self.user.organization_memberships[0].organization_id,
            user_id=self.user.id,
            sku=sku,
            barcode=barcode,
            name="Producto existente",
            category="General",
            cost_price=10,
            sale_price=20,
            stock=stock,
            min_stock=2,
        )
        db.session.add(product)
        db.session.commit()
        return product

    def inventory_html(self):
        response = self.client.get("/products")
        self.assertEqual(response.status_code, 200)
        return response.get_data(as_text=True)

    def import_csv(self, rows, filename="catalog.csv"):
        content = (CSV_HEADER + rows).encode("utf-8")
        return self.client.post(
            "/import-products",
            data={"catalog_file": (io.BytesIO(content), filename)},
            content_type="multipart/form-data",
            follow_redirects=True,
        )

    def test_empty_inventory_hides_catalog_controls(self):
        html = self.inventory_html()

        self.assertIn("Agregar primer producto", html)
        self.assertIn("Importar catálogo", html)
        self.assertNotIn("Buscar producto…", html)
        self.assertNotIn("Seleccionar productos", html)
        self.assertNotIn("Borrar todo el catálogo", html)
        self.assertNotIn('id="catalogo"', html)

    def test_downloadable_template_includes_hardware_store_example(self):
        response = self.client.get("/download-template")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(
            io.BytesIO(response.data),
            read_only=True,
            data_only=True,
        )
        self.assertIn("PRODUCTOS", workbook.sheetnames)
        self.assertIn("EJEMPLO FERRETERÍA", workbook.sheetnames)
        example = workbook["EJEMPLO FERRETERÍA"]
        self.assertEqual(example["A2"].value, "FER-0001")
        self.assertEqual(example["B2"].value, "07501234560001")
        workbook.close()

    def test_inventory_with_products_keeps_catalog_controls(self):
        self.add_product()

        html = self.inventory_html()

        self.assertIn("Buscar producto…", html)
        self.assertIn("Seleccionar productos", html)
        self.assertIn("Borrar todo el catálogo", html)
        self.assertIn("Producto existente", html)
        self.assertIn("<table>", html)
        self.assertIn('data-label="Producto"', html)
        self.assertIn('data-label="Acciones"', html)

    def test_search_without_results_keeps_real_catalog_context(self):
        self.add_product()

        html = self.client.get("/products?q=no-existe").get_data(as_text=True)

        self.assertIn("No encontramos productos", html)
        self.assertIn("Limpiar filtros", html)
        self.assertIn("0 de 1 productos", html)
        self.assertNotIn("Tu catálogo empieza aquí", html)

    def test_large_catalog_is_paginated_without_losing_filters(self):
        organization_id = self.user.organization_memberships[0].organization_id
        db.session.add_all(
            Product(
                organization_id=organization_id,
                user_id=self.user.id,
                sku=f"PAG-{index:03d}",
                barcode=f"7509999{index:05d}",
                name=f"Producto paginado {index:03d}",
                category="Ferretería",
                cost_price=10,
                sale_price=20,
                stock=10,
                min_stock=2,
            )
            for index in range(205)
        )

    def preview_catalog(self, rows, filename="catalog.csv", mapping=None):
        content = (CSV_HEADER + rows).encode("utf-8")
        data = {"catalog_file": (io.BytesIO(content), filename)}
        if mapping is not None:
            data["mapping"] = json.dumps(mapping)
        return self.client.post(
            "/api/products/import/preview",
            data=data,
            content_type="multipart/form-data",
        ), content

    def commit_catalog(self, content, preview, filename="catalog.csv"):
        return self.client.post(
            "/api/products/import/commit",
            data={
                "catalog_file": (io.BytesIO(content), filename),
                "mapping": json.dumps(preview["mapping"]),
                "digest": preview["digest"],
            },
            content_type="multipart/form-data",
        )
        db.session.commit()

        first_page = self.client.get("/products").get_data(as_text=True)
        second_page = self.client.get("/products?page=2").get_data(as_text=True)
        last_page = self.client.get("/products?page=999").get_data(as_text=True)
        filtered_page = self.client.get(
            "/products?q=Producto&page=2"
        ).get_data(as_text=True)

        self.assertEqual(first_page.count("inventory-v3__row-actions"), 100)
        self.assertIn("Producto paginado 000", first_page)
        self.assertNotIn("Producto paginado 150", first_page)
        self.assertIn("Página 1 de 3", first_page)
        self.assertIn("Mostrando 1–100 de 205", first_page)
        self.assertIn("Producto paginado 150", second_page)
        self.assertIn("Página 2 de 3", second_page)
        self.assertIn("q=Producto", filtered_page)
        self.assertIn("Producto paginado 204", last_page)
        self.assertIn("Página 3 de 3", last_page)

    def test_required_sku_is_visible_before_advanced_options(self):
        html = self.inventory_html()

        sku_position = html.index('name="sku"')
        advanced_position = html.index('class="inventory-v2__advanced"')
        self.assertLess(sku_position, advanced_position)

    def test_duplicate_manual_sku_is_rejected_without_server_error(self):
        self.add_product(sku="DUPLICADO")

        response = self.client.post(
            "/products/new",
            data={
                "sku": "DUPLICADO",
                "name": "Segundo producto",
                "cost_price": "5",
                "sale_price": "10",
                "stock": "2",
                "min_stock": "1",
            },
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("Ya existe un producto con ese SKU", response.get_data(as_text=True))
        self.assertEqual(Product.query.filter_by(user_id=self.user.id).count(), 1)

    def test_duplicate_manual_barcode_is_rejected_without_server_error(self):
        self.add_product(barcode="750123")
        response = self.client.post(
            "/products/new",
            data={
                "sku": "SKU-LIBRE", "barcode": "750123", "name": "Segundo producto",
                "cost_price": "5", "sale_price": "10", "stock": "2", "min_stock": "1",
            },
            follow_redirects=True,
        )
        self.assertIn("Ya existe un producto con ese código de barras", response.get_data(as_text=True))
        self.assertEqual(Product.query.filter_by(user_id=self.user.id).count(), 1)

    def test_product_edit_preserves_historical_sale_values(self):
        product = self.add_product()
        sale = Sale(organization_id=self.user.organization_memberships[0].organization_id, user_id=self.user.id, product_id=product.id, quantity=2, unit_price=20, total=40)
        db.session.add(sale)
        db.session.commit()

        response = self.client.post(
            f"/products/{product.id}/edit",
            data={
                "name": "Producto editado", "sku": "EDITADO", "barcode": "8800",
                "category": "Nueva", "supplier": "Proveedor", "cost_price": "12",
                "sale_price": "35", "stock": "9", "min_stock": "3",
            },
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        db.session.refresh(product)
        db.session.refresh(sale)
        self.assertEqual((product.name, product.sale_price, product.stock), ("Producto editado", 35, 9))
        self.assertEqual((sale.unit_price, sale.total), (20, 40))

    def test_product_edit_rejects_other_company_product(self):
        product = self.add_product()
        other = User(email="other@patia.test", company_name="Otra", email_verified=True)
        other.set_password("Password123")
        db.session.add(other)
        db.session.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = other.id

        response = self.client.get(f"/products/{product.id}/edit")

        self.assertEqual(response.status_code, 404)

    def test_product_edit_rejects_duplicate_sku_and_barcode(self):
        product = self.add_product(sku="ORIGINAL", barcode="100")
        self.add_product(sku="OCUPADO", barcode="200")
        payload = {
            "name": "Producto", "sku": "OCUPADO", "barcode": "100",
            "cost_price": "1", "sale_price": "2", "stock": "1", "min_stock": "1",
        }
        self.assertEqual(self.client.post(f"/products/{product.id}/edit", data=payload).status_code, 409)
        payload["sku"] = "LIBRE"
        payload["barcode"] = "200"
        self.assertEqual(self.client.post(f"/products/{product.id}/edit", data=payload).status_code, 409)

    def test_product_edit_validates_numbers_without_server_error(self):
        product = self.add_product()
        response = self.client.post(
            f"/products/{product.id}/edit",
            data={
                "name": "Producto", "sku": "VALIDO", "cost_price": "-1",
                "sale_price": "2", "stock": "1", "min_stock": "1",
            },
        )
        self.assertEqual(response.status_code, 400)

    def test_product_with_sales_is_archived_and_history_is_preserved(self):
        product = self.add_product()
        sale = Sale(
            organization_id=self.user.organization_memberships[0].organization_id,
            user_id=self.user.id,
            product_id=product.id,
            quantity=1,
            unit_price=20,
            total=20,
        )
        db.session.add(sale)
        db.session.commit()

        response = self.client.post(
            f"/products/{product.id}/delete", follow_redirects=True
        )

        html = response.get_data(as_text=True)
        self.assertIn("retirado del catálogo", html)
        self.assertNotIn("Producto existente", html)
        self.assertIsNotNone(db.session.get(Product, product.id))
        self.assertFalse(db.session.get(Product, product.id).is_active)
        self.assertIsNotNone(db.session.get(Sale, sale.id))

    def test_unsold_product_is_physically_deleted_and_empty_state_updates(self):
        product = self.add_product()
        response = self.client.post(
            f"/products/{product.id}/delete",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertIsNone(db.session.get(Product, product.id))
        self.assertIn("Agrega tu primer producto", html)
        self.assertNotIn("Buscar producto…", html)

    def test_archived_product_disappears_from_search_pos_and_dashboard_metrics(self):
        product = self.add_product(stock=5)
        db.session.add(Sale(
            organization_id=self.user.organization_memberships[0].organization_id,
            user_id=self.user.id,
            product_id=product.id,
            quantity=1,
            unit_price=20,
            total=20,
        ))
        db.session.commit()
        self.client.post(f"/products/{product.id}/delete")

        self.assertNotIn("Producto existente", self.client.get("/products?q=Producto").get_data(as_text=True))
        pos_html = self.client.get("/sell").get_data(as_text=True)
        self.assertIn("Agrega inventario antes de vender", pos_html)
        self.assertIn("Producto existente", pos_html)  # El historial de ventas permanece legible.
        dashboard = self.client.get("/").get_data(as_text=True)
        self.assertIn("Agrega tu primer producto", dashboard)
        self.assertEqual(Sale.query.filter_by(user_id=self.user.id).count(), 1)

    def test_adding_archived_sku_reactivates_record_without_losing_sales(self):
        product = self.add_product(sku="ARCHIVO", barcode="700", stock=5)
        sale = Sale(organization_id=self.user.organization_memberships[0].organization_id, user_id=self.user.id, product_id=product.id, quantity=1, unit_price=20, total=20)
        db.session.add(sale)
        db.session.commit()
        self.client.post(f"/products/{product.id}/delete")

        response = self.client.post(
            "/products/new",
            data={
                "sku": "ARCHIVO", "barcode": "700", "name": "Producto reactivado",
                "cost_price": "8", "sale_price": "16", "stock": "4", "min_stock": "1",
            },
            follow_redirects=True,
        )

        db.session.refresh(product)
        self.assertTrue(product.is_active)
        self.assertEqual(product.name, "Producto reactivado")
        self.assertIn("Producto reactivado", response.get_data(as_text=True))
        self.assertIsNotNone(db.session.get(Sale, sale.id))

    def test_delete_all_preserves_products_with_sales(self):
        protected = self.add_product(sku="CON-VENTA")
        removable = self.add_product(sku="SIN-VENTA", barcode="7502")
        db.session.add(Sale(
            organization_id=self.user.organization_memberships[0].organization_id,
            user_id=self.user.id,
            product_id=protected.id,
            quantity=1,
            unit_price=20,
            total=20,
        ))
        db.session.commit()

        response = self.client.post("/products/delete-all", follow_redirects=True)

        self.assertIn("retiramos 1", response.get_data(as_text=True))
        self.assertIsNotNone(db.session.get(Product, protected.id))
        self.assertFalse(db.session.get(Product, protected.id).is_active)
        self.assertIsNone(db.session.get(Product, removable.id))
        self.assertEqual(Sale.query.filter_by(user_id=self.user.id).count(), 1)

    def test_manual_form_keeps_names_and_numeric_validation(self):
        html = self.inventory_html()
        for name in (
            "name", "cost_price", "sale_price", "stock", "sku",
            "barcode", "category", "supplier", "min_stock",
        ):
            self.assertRegex(html, rf'<input[^>]*name="{name}"')

        self.assertRegex(html, r'<input[^>]*name="name"[^>]*required')
        for name, step in (("cost_price", "0.01"), ("sale_price", "0.01"),
                           ("stock", "1"), ("min_stock", "1")):
            tag = re.search(rf'<input[^>]*name="{name}"[^>]*>', html).group(0)
            self.assertIn('min="0"', tag)
            self.assertIn(f'step="{step}"', tag)
        self.assertIn("Opciones avanzadas", html)

    def test_import_creates_new_product(self):
        response = self.import_csv("NUEVO,7502,Café,Abarrotes,,12,24,8,3\n")

        product = Product.query.filter_by(user_id=self.user.id, sku="NUEVO").one()
        self.assertEqual(product.stock, 8)
        self.assertIn("1 creados, 0 actualizados, 0 omitidos y 0 errores", response.get_data(as_text=True))

    def test_sku_match_adds_stock_and_updates_values(self):
        product = self.add_product(stock=5)
        self.import_csv("EXISTE,9999,Actualizado,General,,15,30,3,4\n")

        db.session.refresh(product)
        self.assertEqual(product.stock, 8)
        self.assertEqual(product.cost_price, 15)
        self.assertEqual(product.sale_price, 30)
        self.assertEqual(product.min_stock, 4)
        self.assertEqual(product.barcode, "9999")

    def test_barcode_match_replaces_stock_and_prices(self):
        product = self.add_product(sku="ORIGINAL", barcode="7503", stock=9)
        self.import_csv("DISTINTO,7503,Otro nombre,General,,7,18,2,1\n")

        db.session.refresh(product)
        self.assertEqual(product.sku, "ORIGINAL")
        self.assertEqual(product.stock, 2)
        self.assertEqual(product.cost_price, 7)
        self.assertEqual(product.sale_price, 18)
        self.assertEqual(product.min_stock, 1)

    def test_invalid_row_is_counted_without_internal_error_details(self):
        response = self.import_csv("MALO,7504,Producto,General,,no-es-numero,20,2,1\n")
        html = response.get_data(as_text=True)

        self.assertIn("0 creados, 0 actualizados, 0 omitidos y 1 errores", html)
        self.assertNotIn("could not convert", html)
        self.assertEqual(Product.query.filter_by(user_id=self.user.id).count(), 0)

    def test_fractional_or_non_finite_inventory_values_are_rejected(self):
        response = self.import_csv(
            "FRACCION,7505,Producto,General,,10,20,1.5,1\n"
            "INFINITO,7506,Producto,General,,inf,20,2,1\n"
        )

        self.assertIn("0 creados, 0 actualizados, 0 omitidos y 2 errores", response.get_data(as_text=True))
        self.assertEqual(Product.query.filter_by(user_id=self.user.id).count(), 0)

    def test_csv_error_log_reports_the_actual_file_row(self):
        with self.assertLogs("app", level="WARNING") as captured:
            self.import_csv(
                "VALIDO,7507,Producto,General,,10,20,2,1\n"
                "MALO,7508,Producto,General,,texto,20,2,1\n"
            )

        self.assertTrue(any("fila 3" in message for message in captured.output))

    def test_missing_essential_columns_rejects_file(self):
        response = self.client.post(
            "/import-products",
            data={"catalog_file": (io.BytesIO(b"SKU,Nombre del producto\nA,Producto\n"), "bad.csv")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)

        self.assertIn("no contiene las columnas obligatorias", html)
        self.assertEqual(Product.query.filter_by(user_id=self.user.id).count(), 0)

    def test_import_summary_counts_every_result(self):
        self.add_product(sku="POR-SKU", barcode="111", stock=4)
        self.add_product(sku="POR-CODIGO", barcode="222", stock=7)
        rows = (
            "NUEVO,333,Nuevo,General,,4,8,2,1\n"
            "POR-SKU,444,SKU,General,,5,10,3,2\n"
            "OTRO,222,Codigo,General,,6,12,1,2\n"
            ",,,,,,,,\n"
            "ERROR,555,Error,General,,texto,10,1,1\n"
        )

        response = self.import_csv(rows)

        self.assertIn(
            "1 creados, 2 actualizados, 1 omitidos y 1 errores",
            response.get_data(as_text=True),
        )

    def test_imports_one_thousand_products_in_one_safe_batch(self):
        rows = "".join(
            f"FER-{index:04d},{7500000000000 + index},Producto {index},"
            f"Ferretería,Proveedor Uno,1.25,2.50,10,2\n"
            for index in range(1000)
        )

        started_at = time.perf_counter()
        response = self.import_csv(rows)
        elapsed = time.perf_counter() - started_at

        self.assertEqual(response.status_code, 200)
        self.assertIn("1000 creados", response.get_data(as_text=True))
        self.assertEqual(
            Product.query.filter_by(
                organization_id=self.user.organization_memberships[0].organization_id
            ).count(),
            1000,
        )
        self.assertLess(elapsed, 30)

    def test_professional_preview_maps_headers_and_preserves_leading_zeroes(self):
        rows = (
            'PIN-001,0007500123456,"Pintura acrílica blanca 19 L",'
            'Pinturas,Distribuidora Centro,"$1,250.50","2.499,90",'
            "12,3\n"
        )
        response, content = self.preview_catalog(rows)
        self.assertEqual(response.status_code, 200)
        preview = response.get_json()
        self.assertTrue(preview["ready"])
        self.assertEqual(preview["summary"]["new"], 1)
        self.assertEqual(preview["rows"][0]["barcode"], "0007500123456")
        self.assertEqual(preview["rows"][0]["cost_price"], "1250.50")
        self.assertEqual(preview["rows"][0]["sale_price"], "2499.90")
        committed = self.commit_catalog(content, preview)
        self.assertEqual(committed.status_code, 200)
        product = Product.query.filter_by(sku="PIN-001").one()
        self.assertEqual(product.barcode, "0007500123456")
        self.assertEqual(product.supplier, "Distribuidora Centro")

    def test_professional_import_is_idempotent_and_does_not_duplicate_stock(self):
        rows = "IDEM-001,0001,Martillo,Herrajes,,50,90,14,3\n"
        preview_response, content = self.preview_catalog(rows)
        preview = preview_response.get_json()
        self.assertEqual(self.commit_catalog(content, preview).status_code, 200)
        first_movement_count = InventoryMovement.query.count()
        repeated_response, repeated_content = self.preview_catalog(rows)
        repeated = repeated_response.get_json()
        self.assertEqual(repeated["summary"]["updated"], 1)
        self.assertEqual(
            self.commit_catalog(repeated_content, repeated).status_code,
            200,
        )
        self.assertEqual(Product.query.filter_by(sku="IDEM-001").count(), 1)
        self.assertEqual(Product.query.filter_by(sku="IDEM-001").one().stock, 14)
        self.assertEqual(InventoryMovement.query.count(), first_movement_count)

    def test_professional_preview_imports_valid_rows_and_separates_duplicates(self):
        rows = (
            "DUP-001,0001,Taladro,Herramientas,,500,900,2,1\n"
            "DUP-001,0002,Segundo,Herramientas,,100,200,3,1\n"
        )
        response, content = self.preview_catalog(rows)
        preview = response.get_json()
        self.assertTrue(preview["ready"])
        self.assertEqual(preview["summary"]["duplicates"], 1)
        committed = self.commit_catalog(content, preview)
        self.assertEqual(committed.status_code, 200)
        self.assertEqual(Product.query.count(), 1)
        self.assertEqual(committed.get_json()["summary"]["errors"], 1)
        self.assertIn("repetido", preview["errors"][0]["message"])

    def test_existing_non_patia_file_is_detected_and_missing_sku_is_generated(self):
        content = (
            "Catálogo Ferretería Puebla;;;;\n"
            "Artículo;Código Universal;Familia;Costo Unitario;Precio Público;Existencia;Existencia Mínima\n"
            "Martillo carpintero;0007500999001;Herramientas;80.50;149.90;25;4\n"
        ).encode("utf-8")
        response = self.client.post(
            "/api/products/import/preview",
            data={"catalog_file": (io.BytesIO(content), "mi_inventario.csv")},
            content_type="multipart/form-data",
        )
        preview = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(preview["ready"])
        self.assertEqual(preview["mapping"]["name"], "Artículo")
        self.assertEqual(preview["mapping"]["barcode"], "Código Universal")
        self.assertEqual(preview["mapping"]["sale_price"], "Precio Público")
        self.assertTrue(preview["rows"][0]["sku"].startswith("IMP-"))
        committed = self.commit_catalog(
            content, preview, filename="mi_inventario.csv"
        )
        self.assertEqual(committed.status_code, 200)
        product = Product.query.one()
        self.assertEqual(product.barcode, "0007500999001")
        self.assertEqual(product.name, "Martillo carpintero")

    def test_external_erp_headers_map_identifiers_and_reorder_point(self):
        existing = self.add_product(
            sku="OLD-000001",
            barcode="7500000000001",
            stock=2,
        )
        content = (
            "CATALOGO EXPORTADO DEL SISTEMA ANTERIOR;;;;;;;;\n"
            "No. articulo;GTIN / EAN;Descripcion comercial;"
            "Familia de articulos;Marca proveedor;Precio compra;"
            "Precio publico;Existencia;Punto de reorden\n"
            "OLD-000001;7500000000001;Martillo actualizado;"
            "Herramientas;Proveedor Norte;45.20;99.90;18;6\n"
        ).encode("utf-8")

        response = self.client.post(
            "/api/products/import/preview",
            data={"catalog_file": (io.BytesIO(content), "exportacion-erp.csv")},
            content_type="multipart/form-data",
        )
        preview = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(preview["mapping"]["sku"], "No. articulo")
        self.assertEqual(preview["mapping"]["barcode"], "GTIN / EAN")
        self.assertEqual(
            preview["mapping"]["name"], "Descripcion comercial"
        )
        self.assertEqual(
            preview["mapping"]["min_stock"], "Punto de reorden"
        )
        self.assertEqual(preview["summary"]["updated"], 1)
        self.assertEqual(
            self.commit_catalog(
                content, preview, filename="exportacion-erp.csv"
            ).status_code,
            200,
        )
        db.session.refresh(existing)
        self.assertEqual(existing.stock, 18)
        self.assertEqual(existing.min_stock, 6)

    def test_professional_import_is_isolated_between_organizations(self):
        rows = "SHARED-001,0007,Llave inglesa,Herramientas,,70,130,5,2\n"
        first_preview, first_content = self.preview_catalog(rows)
        self.assertEqual(
            self.commit_catalog(first_content, first_preview.get_json()).status_code,
            200,
        )
        other = User(
            email="other-import@patia.test",
            company_name="Otra ferretería",
            email_verified=True,
        )
        other.set_password("Password123")
        db.session.add(other)
        db.session.flush()
        ensure_owner_organization(other)
        db.session.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = other.id
        second_preview, second_content = self.preview_catalog(rows)
        self.assertEqual(second_preview.get_json()["summary"]["new"], 1)
        self.assertEqual(
            self.commit_catalog(
                second_content, second_preview.get_json()
            ).status_code,
            200,
        )
        self.assertEqual(Product.query.filter_by(sku="SHARED-001").count(), 2)

    def test_professional_import_handles_five_thousand_products_in_one_transaction(self):
        rows = "".join(
            f"BIG-{index:05d},{index:013d},Producto ferretería {index},"
            f"Herramientas,Proveedor Mayorista,10.25,18.50,20,4\n"
            for index in range(5000)
        )
        started = time.perf_counter()
        preview_response, content = self.preview_catalog(rows)
        preview = preview_response.get_json()
        self.assertEqual(preview["summary"]["valid"], 5000)
        committed = self.commit_catalog(content, preview)
        elapsed = time.perf_counter() - started
        self.assertEqual(committed.status_code, 200)
        self.assertEqual(Product.query.count(), 5000)
        self.assertEqual(InventoryMovement.query.count(), 5000)
        self.assertLess(elapsed, 45)

    def test_import_form_blocks_double_submission(self):
        html = self.inventory_html()

        self.assertIn("let importSubmitting = false", html)
        self.assertIn("if (importSubmitting)", html)
        self.assertIn("importButton.disabled = true", html)
        self.assertIn(r"Validando cat\u00e1logo", html)
        self.assertIn("/api/products/import/preview", html)
        self.assertIn("/api/products/import/commit", html)


if __name__ == "__main__":
    unittest.main()
